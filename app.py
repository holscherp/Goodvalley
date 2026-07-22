import os, re as _re_app
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from sqlalchemy.exc import IntegrityError

from db import db

_OT_SUFFIX_RE = _re_app.compile(r'^(.+)-(\d{1,2})$')

def _ot_base(ot):
    """Return the base OT, stripping trailing -1 / -2 style line suffixes.
    e.g. '790-0001-1' → '790-0001', '807-0005' → '807-0005' (unchanged)."""
    m = _OT_SUFFIX_RE.match(ot or '')
    if m:
        # Only strip if the prefix itself looks like a real OT (contains a dash)
        prefix = m.group(1)
        if '-' in prefix:
            return prefix
    return ot

PWAREHOUSE_URL  = os.environ.get('PWAREHOUSE_URL',  'http://190.211.168.247:8077')
PWAREHOUSE_RUT  = os.environ.get('PWAREHOUSE_RUT',  '')
PWAREHOUSE_PASS = os.environ.get('PWAREHOUSE_PASS', '')

DRYING_MAP = {
    'cancha': 'cancha', 'cancha de sol': 'cancha', 'sol': 'cancha',
    'campo': 'cancha', 'field': 'cancha', 'field drying': 'cancha',
    'horno': 'horno', 'oven': 'horno', 'oven drying': 'horno',
    'termino secado': 'termino_secado', 'término secado': 'termino_secado',
    'termino_secado': 'termino_secado', 'term. secado': 'termino_secado',
    'term.secado': 'termino_secado',
}


_SYNC_POPUP = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Sync pWarehouse</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0 }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #0f0f1a; color: #e0d4f7;
    display: flex; flex-direction: column; height: 100vh; overflow: hidden;
  }
  #header {
    background: #3b0764; padding: 12px 16px;
    font-size: 14px; font-weight: 600; flex-shrink: 0;
    display: flex; align-items: center; gap: 10px;
  }
  #spinner {
    width: 14px; height: 14px; border: 2px solid rgba(255,255,255,.3);
    border-top-color: #fff; border-radius: 50%;
    animation: spin .7s linear infinite; flex-shrink: 0;
  }
  @keyframes spin { to { transform: rotate(360deg) } }
  #log {
    flex: 1; overflow-y: auto; padding: 12px 16px;
    font-family: 'SF Mono', monospace; font-size: 12px;
    line-height: 1.65; white-space: pre-wrap; color: #d4f0c0;
  }
  #footer {
    padding: 10px 16px; background: #1a1a2e; flex-shrink: 0;
    font-size: 13px; font-weight: 600; min-height: 38px;
  }
  .ok  { color: #6fcf97 }
  .err { color: #eb5757 }
</style>
</head>
<body>
<div id="header">
  <div id="spinner"></div>
  <span id="title">Sincronizando pWarehouse…</span>
</div>
<div id="log"></div>
<div id="footer"></div>
<script>
(function() {
  const log     = document.getElementById('log');
  const footer  = document.getElementById('footer');
  const title   = document.getElementById('title');
  const spinner = document.getElementById('spinner');

  let offset = 0;
  let timer  = null;

  function finish(ok) {
    clearInterval(timer);
    spinner.style.display = 'none';
    if (ok) {
      title.textContent = '\\u2713 Sincronizaci\\u00f3n completada';
      footer.innerHTML  = '<span class="ok">\\u2713 Listo \\u2014 cerrando en 2 segundos\\u2026</span>';
      setTimeout(function() {
        try { if (window.opener) window.opener.postMessage('gv_sync_done', '*'); } catch(_) {}
        window.close();
      }, 2000);
    } else {
      title.textContent = 'Error al sincronizar';
      footer.innerHTML  = '<span class="err">\\u2717 Error \\u2014 revis\\u00e1 el log arriba.</span>';
    }
  }

  function poll(jobId) {
    fetch('/sync/status/' + jobId + '?offset=' + offset)
      .then(function(r) { return r.json(); })
      .then(function(d) {
        (d.lines || []).forEach(function(line) {
          log.textContent += line + '\\n';
          log.scrollTop = log.scrollHeight;
        });
        offset = d.offset;
        if (d.done) finish(d.returncode === 0);
      })
      .catch(function(e) { console.warn('poll error (retrying):', e); });
  }

  fetch('/sync/start', { method: 'POST' })
    .then(function(r) { return r.json(); })
    .then(function(d) {
      if (!d.job_id) { finish(false); return; }
      timer = setInterval(function() { poll(d.job_id); }, 2000);
    })
    .catch(function() {
      spinner.style.display = 'none';
      title.textContent = 'Error de conexi\\u00f3n';
      footer.innerHTML  = '<span class="err">\\u2717 No se pudo iniciar la sincronizaci\\u00f3n.</span>';
    });
})();
</script>
</body>
</html>"""


def create_app():
    app = Flask(__name__)

    db_url = os.environ.get('DATABASE_URL', 'postgresql://localhost/goodvalley')
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)

    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-change-me')

    db.init_app(app)

    with app.app_context():
        from models import Bin, Order, OrderLine, Allocation, Excedente, YieldOverride, Proceso, HistoricoMovimiento, OrdenDeVenta, Pallet, AppSetting  # noqa: F401
        _pre_migrate(db)
        db.create_all()

        _migrate(db)

        from models import load_yield_overrides
        load_yield_overrides()

    # ── Routes ────────────────────────────────────────────────────────────────

    @app.route('/')
    def index():
        from models import Bin, Order, Allocation
        from sqlalchemy import func

        total   = Bin.query.count()
        avail   = Bin.query.filter_by(status='available').count()
        kg      = db.session.query(func.sum(Bin.weight_kg)).filter_by(status='available').scalar() or 0
        n_open  = Order.query.filter(Order.status.in_(['open', 'confirmed'])).count()
        alloc_n = Allocation.query.count()

        # Summary by caliber + drying (available bins only)
        rows = (
            db.session.query(
                Bin.caliber, Bin.drying,
                func.count(Bin.id).label('cnt'),
                func.sum(Bin.weight_kg).label('kg'),
            )
            .filter_by(status='available')
            .filter(Bin.caliber.isnot(None))
            .group_by(Bin.caliber, Bin.drying)
            .order_by(Bin.caliber, Bin.drying)
            .all()
        )

        from models import DRYING_LABELS
        return render_template('index.html',
            total=total, avail=avail, kg=round(kg, 1),
            n_open=n_open, alloc_n=alloc_n, summary_rows=rows,
            DRYING_LABELS=DRYING_LABELS,
        )

    # ── Sync ──────────────────────────────────────────────────────────────────

    @app.route('/sync-popup')
    def sync_popup():
        from flask import make_response
        resp = make_response(_SYNC_POPUP)
        resp.headers['Content-Type'] = 'text/html; charset=utf-8'
        return resp

    @app.route('/sync-full-popup')
    def sync_full_popup():
        from flask import make_response
        html = _SYNC_POPUP.replace("fetch('/sync/start',", "fetch('/sync/full/start',")
        resp = make_response(html)
        resp.headers['Content-Type'] = 'text/html; charset=utf-8'
        return resp

    @app.route('/sync/start', methods=['POST'])
    def sync_start():
        import subprocess, sys, threading, uuid as _uuid
        from pathlib import Path as _Path
        from flask import jsonify, current_app as _ca
        from models import Bin

        job_id      = _uuid.uuid4().hex[:8]
        log_path    = _Path(f'/tmp/gv_job_{job_id}.log')
        status_path = _Path(f'/tmp/gv_job_{job_id}.status')
        bins_path   = _Path(f'/tmp/gv_bins_{job_id}.json')
        pallets_path = _Path(f'/tmp/gv_pallets_{job_id}.json')
        scraper     = _Path(__file__).parent / 'scrape_pwarehouse.py'

        log_path.write_text('')
        status_path.write_text('running')

        _app = _ca._get_current_object()

        def _temporada(t):
            if len(t) >= 8:
                try:
                    p = int(t[:2])
                    if 18 <= p <= 35:
                        return str(2000 + p)
                except ValueError:
                    pass
            return None

        def run():
            env = {**os.environ, 'GV_NO_UPLOAD': '1', 'GV_OUTPUT': str(bins_path), 'GV_PALLETS_OUT': str(pallets_path)}
            proc = subprocess.Popen(
                [sys.executable, str(scraper)],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, env=env,
            )
            with open(log_path, 'a') as lf:
                for line in proc.stdout:
                    lf.write(line.rstrip() + '\n')
                    lf.flush()
            proc.wait()

            if proc.returncode != 0 or not bins_path.exists():
                status_path.write_text(f'done:{proc.returncode or 1}')
                return

            import json as _json
            try:
                bins_data = _json.loads(bins_path.read_text())
                with open(log_path, 'a') as lf:
                    lf.write('► Importando bins a la base de datos...\n')
                    lf.flush()

                with _app.app_context():
                    existing_map = {
                        row[0]: row[1] for row in
                        db.session.query(Bin.bin_identifier, Bin.id)
                        .filter(Bin.status == 'available').all()
                    }
                    all_ids = {row[0] for row in db.session.query(Bin.bin_identifier).all()}

                    added = updated = skipped = 0
                    new_batch = []

                    for b in bins_data:
                        bid = str(b.get('bin_identifier', '')).strip()
                        if not bid:
                            skipped += 1; continue
                        drying = b.get('drying') or ''
                        if drying not in ('cancha', 'horno', 'termino_secado'):
                            skipped += 1; continue
                        weight = float(b.get('weight_kg') or 0)

                        if bid in existing_map:
                            db.session.query(Bin).filter_by(id=existing_map[bid]).update({
                                'weight_kg': weight,
                                'humedad': b.get('humedad'),
                                'caliber': b.get('caliber') or '',
                                'u_lb': b.get('u_lb'),
                                'drying': drying,
                                'producto': b.get('producto') or '',
                                'contenedor': b.get('contenedor') or '',
                                'producer_name': b.get('producer_name') or '',
                                'temporada': b.get('temporada') or _temporada(bid),
                            }, synchronize_session=False)
                            updated += 1
                        elif bid not in all_ids:
                            new_batch.append(Bin(
                                bin_identifier=bid,
                                producto=b.get('producto') or '',
                                caliber=b.get('caliber') or '',
                                u_lb=b.get('u_lb'),
                                drying=drying,
                                weight_kg=weight,
                                humedad=b.get('humedad'),
                                contenedor=b.get('contenedor') or '',
                                producer_name=b.get('producer_name') or '',
                                temporada=b.get('temporada') or _temporada(bid),
                                status='available',
                            ))
                            all_ids.add(bid)
                            added += 1
                            if len(new_batch) >= 500:
                                db.session.bulk_save_objects(new_batch)
                                db.session.commit()
                                new_batch = []
                        else:
                            skipped += 1

                    if new_batch:
                        db.session.bulk_save_objects(new_batch)
                    db.session.commit()

                with open(log_path, 'a') as lf:
                    lf.write(f'✓ {added} nuevos, {updated} actualizados, {skipped} omitidos.\n')
                    lf.flush()

                # ── Import pallets ─────────────────────────────────────────
                if pallets_path.exists():
                    try:
                        from models import Pallet as _Pallet, HistoricoMovimiento as _HM
                        from datetime import datetime as _dt
                        pallets_data = _json.loads(pallets_path.read_text())
                        with open(log_path, 'a') as lf:
                            lf.write(f'► Importando {len(pallets_data)} pallets...\n')
                            lf.flush()
                        with _app.app_context():
                            shipped_ots = {
                                r[0] for r in db.session.query(_HM.ot)
                                .filter(_HM.movimiento == 'EMBARQUE').all()
                            }
                            existing_p = {row[0] for row in db.session.query(_Pallet.tarja).all()}
                            added_p = updated_p = 0
                            for p in pallets_data:
                                tarja = str(p.get('tarja') or '').strip()
                                if not tarja:
                                    continue
                                fields = {
                                    'ot':               str(p.get('ot') or '').strip(),
                                    'customer':         p.get('customer') or '',
                                    'caliber':          p.get('caliber') or '',
                                    'drying':           p.get('drying') or '',
                                    'weight_kg':        float(p.get('weight_kg') or 0),
                                    'producto':         p.get('producto') or '',
                                    'unidades':         p.get('unidades'),
                                    'pallet_estado_ot': p.get('pallet_estado_ot'),
                                    's_pallet_clase':   p.get('s_pallet_clase'),
                                    'synced_at':        _dt.utcnow(),
                                }
                                if tarja in existing_p:
                                    db.session.query(_Pallet).filter_by(tarja=tarja).update(
                                        fields, synchronize_session=False)
                                    updated_p += 1
                                else:
                                    db.session.add(_Pallet(tarja=tarja, **fields))
                                    existing_p.add(tarja)
                                    added_p += 1
                            db.session.commit()
                        with open(log_path, 'a') as lf:
                            lf.write(f'✓ Pallets: {added_p} nuevos, {updated_p} actualizados.\n')
                            lf.flush()
                    except Exception as _ep:
                        with open(log_path, 'a') as lf:
                            lf.write(f'⚠ Error importando pallets: {_ep}\n')
                            lf.flush()

                status_path.write_text('done:0')

            except Exception as e:
                with open(log_path, 'a') as lf:
                    lf.write(f'✗ Error importando: {e}\n')
                    lf.flush()
                status_path.write_text('done:1')

        threading.Thread(target=run, daemon=True).start()
        return jsonify({'job_id': job_id})

    @app.route('/sync/status/<job_id>')
    def sync_status(job_id):
        from pathlib import Path as _Path
        from flask import jsonify

        log_path    = _Path(f'/tmp/gv_job_{job_id}.log')
        status_path = _Path(f'/tmp/gv_job_{job_id}.status')

        offset = request.args.get('offset', 0, type=int)

        lines = []
        if log_path.exists():
            all_lines = [l for l in log_path.read_text().split('\n') if l]
            lines = all_lines[offset:]

        raw_status  = status_path.read_text().strip() if status_path.exists() else 'running'
        done        = raw_status.startswith('done:')
        returncode  = int(raw_status.replace('done:', '')) if done else None

        return jsonify({
            'lines': lines,
            'offset': offset + len(lines),
            'done': done,
            'returncode': returncode,
        })

    @app.route('/sync/full/start', methods=['POST'])
    def sync_full_start():
        import subprocess, sys, threading, uuid as _uuid
        from pathlib import Path as _Path
        from flask import jsonify, current_app as _ca
        from models import Bin, Order, Allocation

        job_id        = _uuid.uuid4().hex[:8]
        log_path      = _Path(f'/tmp/gv_job_{job_id}.log')
        status_path   = _Path(f'/tmp/gv_job_{job_id}.status')
        bins_path     = _Path(f'/tmp/gv_bins_{job_id}.json')
        pallets_path  = _Path(f'/tmp/gv_pallets_{job_id}.json')
        procesos_path = _Path(f'/tmp/gv_procesos_{job_id}.json')
        scraper       = _Path(__file__).parent / 'scrape_full.py'

        log_path.write_text('')
        status_path.write_text('running')

        _app = _ca._get_current_object()

        def _infer_temporada(tarja_str):
            if len(tarja_str) >= 8:
                try:
                    p = int(tarja_str[:2])
                    if 18 <= p <= 35:
                        return str(2000 + p)
                except ValueError:
                    pass
            return None

        def run():
            import json as _json

            env = {
                **os.environ,
                'GV_NO_UPLOAD':   '1',
                'GV_BINS_OUT':    str(bins_path),
                'GV_PALLETS_OUT': str(pallets_path),
                'GV_PROCESOS_OUT': str(procesos_path),
            }
            proc = subprocess.Popen(
                [sys.executable, str(scraper)],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, env=env,
            )
            with open(log_path, 'a') as lf:
                for line in proc.stdout:
                    lf.write(line.rstrip() + '\n')
                    lf.flush()
            proc.wait()

            if proc.returncode != 0:
                status_path.write_text(f'done:{proc.returncode or 1}')
                return

            # ── Import bins ───────────────────────────────────────────────────
            if bins_path.exists():
                try:
                    bins_data = _json.loads(bins_path.read_text())
                    with open(log_path, 'a') as lf:
                        lf.write('► Importando bins a la base de datos...\n')
                        lf.flush()
                    with _app.app_context():
                        from models import OrderLine, Order, Allocation
                        existing_map = {
                            row[0]: row[1] for row in
                            db.session.query(Bin.bin_identifier, Bin.id)
                            .filter(Bin.status == 'available').all()
                        }
                        all_ids = {row[0] for row in db.session.query(Bin.bin_identifier).all()}
                        added = updated = skipped = 0
                        new_batch = []
                        to_allocate = []
                        for b in bins_data:
                            bid = str(b.get('bin_identifier', '')).strip()
                            if not bid: skipped += 1; continue
                            drying = b.get('drying') or ''
                            if drying not in ('cancha', 'horno', 'termino_secado'):
                                skipped += 1; continue
                            weight = float(b.get('weight_kg') or 0)
                            if bid in existing_map:
                                db.session.query(Bin).filter_by(id=existing_map[bid]).update({
                                    'weight_kg': weight,
                                    'humedad': b.get('humedad'),
                                    'caliber': b.get('caliber') or '',
                                    'u_lb': b.get('u_lb'),
                                    'drying': drying,
                                    'producto': b.get('producto') or '',
                                    'contenedor': b.get('contenedor') or '',
                                    'producer_name': b.get('producer_name') or '',
                                    'temporada': b.get('temporada') or _infer_temporada(bid),
                                }, synchronize_session=False)
                                updated += 1
                            elif bid not in all_ids:
                                new_batch.append(Bin(
                                    bin_identifier=bid,
                                    producto=b.get('producto') or '',
                                    caliber=b.get('caliber') or '',
                                    u_lb=b.get('u_lb'),
                                    drying=drying,
                                    weight_kg=weight,
                                    humedad=b.get('humedad'),
                                    contenedor=b.get('contenedor') or '',
                                    producer_name=b.get('producer_name') or '',
                                    temporada=b.get('temporada') or _infer_temporada(bid),
                                    status='available',
                                ))
                                all_ids.add(bid)
                                added += 1
                                ot = b.get('ot')
                                if ot:
                                    to_allocate.append((bid, str(ot).strip()))
                                if len(new_batch) >= 500:
                                    db.session.bulk_save_objects(new_batch)
                                    db.session.commit()
                                    new_batch = []
                            else:
                                skipped += 1
                        if new_batch:
                            db.session.bulk_save_objects(new_batch)
                        db.session.commit()

                        # Auto-allocate new bins via OT → order line match
                        alloc_count = 0
                        if to_allocate:
                            ot_line_map = {}
                            for ln in (OrderLine.query
                                       .join(Order, OrderLine.order_id == Order.id)
                                       .filter(Order.status.in_(['open', 'confirmed']))
                                       .filter(OrderLine.notes.isnot(None))
                                       .all()):
                                for part in (ln.notes or '').split('·'):
                                    part = part.strip()
                                    if part.startswith('OT '):
                                        ot_key = part[3:].strip()
                                        ot_line_map.setdefault(ot_key, ln)
                            for bid, ot in to_allocate:
                                line = ot_line_map.get(ot)
                                if not line:
                                    continue
                                b_obj = Bin.query.filter_by(
                                    bin_identifier=bid, status='available').first()
                                if not b_obj:
                                    continue
                                if Allocation.query.filter_by(bin_id=b_obj.id).first():
                                    continue
                                db.session.add(Allocation(
                                    order_id=line.order_id, line_id=line.id, bin_id=b_obj.id))
                                b_obj.status = 'allocated'
                                alloc_count += 1
                            db.session.commit()

                    with open(log_path, 'a') as lf:
                        lf.write(
                            f'✓ Bins: {added} nuevos, {updated} actualizados, '
                            f'{skipped} omitidos, {alloc_count} auto-asignados.\n'
                        )
                        lf.flush()
                except Exception as e:
                    with open(log_path, 'a') as lf:
                        lf.write(f'✗ Error importando bins: {e}\n')
                        lf.flush()

            # ── Import pallets ────────────────────────────────────────────────
            if pallets_path.exists():
                try:
                    from models import Pallet as _Pallet
                    pallets_data = _json.loads(pallets_path.read_text())
                    with open(log_path, 'a') as lf:
                        lf.write(f'► Importando {len(pallets_data)} pallets...\n')
                        lf.flush()
                    with _app.app_context():
                        import datetime as _dt
                        existing_p = {row[0] for row in db.session.query(_Pallet.tarja).all()}
                        added_p = updated_p = 0
                        for p in pallets_data:
                            tarja = str(p.get('tarja') or '').strip()
                            if not tarja:
                                continue
                            fields = {
                                'ot':               str(p.get('ot') or '').strip() or None,
                                'customer':         str(p.get('customer') or '').strip() or None,
                                'caliber':          str(p.get('caliber') or '').strip() or None,
                                'drying':           p.get('drying') or None,
                                'product_type':     p.get('product_type') or None,
                                'weight_kg':        float(p.get('weight_kg') or 0),
                                'producto':         str(p.get('producto') or '').strip() or None,
                                'temporada':        str(p.get('temporada') or '').strip() or None,
                                'pallet_estado_ot': p.get('pallet_estado_ot') or None,
                                's_pallet_clase':   p.get('s_pallet_clase') or None,
                                'unidades':         int(p['unidades']) if p.get('unidades') else None,
                                'synced_at':        _dt.datetime.utcnow(),
                            }
                            if tarja in existing_p:
                                db.session.query(_Pallet).filter_by(tarja=tarja).update(
                                    fields, synchronize_session=False)
                                updated_p += 1
                            else:
                                db.session.add(_Pallet(tarja=tarja, **fields))
                                existing_p.add(tarja)
                                added_p += 1
                        db.session.commit()
                    with open(log_path, 'a') as lf:
                        lf.write(f'✓ Pallets: {added_p} nuevos, {updated_p} actualizados.\n')
                        lf.flush()
                except Exception as _ep:
                    with open(log_path, 'a') as lf:
                        lf.write(f'⚠ Error importando pallets: {_ep}\n')
                        lf.flush()

            status_path.write_text('done:0')

        threading.Thread(target=run, daemon=True).start()
        return jsonify({'job_id': job_id})

    @app.route('/sync', methods=['POST'])
    def sync():
        import json as _json, os as _os, re as _re
        from models import Bin, _parse_producto

        _DRYING_MAP = dict(DRYING_MAP)

        def _parse_js(text):
            try:
                return _json.loads(text)
            except Exception:
                pass
            return _json.loads(_re.sub(r'(?<!["\w])(\w+)\s*:', r'"\1":', text))

        def _rv(row, i):
            return row.get(i) or row.get(str(i))

        def _temporada(t):
            if len(t) >= 8:
                try:
                    p = int(t[:2])
                    if 18 <= p <= 35:
                        return str(2000 + p)
                except ValueError:
                    pass
            return None

        def _do_import(bins_data):
            added = skipped = 0
            for b in bins_data:
                bid = str(b.get('bin_identifier', '')).strip()
                if not bid:
                    skipped += 1
                    continue
                exists = Bin.query.filter_by(bin_identifier=bid).first()
                if exists:
                    skipped += 1
                    continue
                drying = b.get('drying') or b.get('drying_method', '')
                if drying not in ('cancha', 'horno', 'termino_secado'):
                    skipped += 1
                    continue
                db.session.add(Bin(
                    bin_identifier=bid,
                    producto=b.get('producto') or b.get('product') or '',
                    caliber=b.get('caliber') or b.get('caliber_str') or '',
                    drying=drying,
                    weight_kg=float(b.get('weight_kg') or 0),
                    humedad=b.get('humedad') or b.get('humidity'),
                    contenedor=b.get('contenedor') or b.get('container', ''),
                    producer_name=b.get('producer_name', ''),
                    temporada=b.get('temporada') or _temporada(bid),
                    status='available',
                ))
                added += 1
            db.session.commit()
            return added, skipped

        live = request.form.get('live') == '1'

        try:
            if not live:
                dump_path = _os.path.join(_os.path.dirname(__file__), 'data_dump', 'bins.json')
                if not _os.path.exists(dump_path):
                    flash('data_dump/bins.json no encontrado. Coloca el archivo y vuelve a hacer deploy.', 'err')
                    return redirect(url_for('index'))
                with open(dump_path) as f:
                    bins_data = _json.load(f)
                added, skipped = _do_import(bins_data)
                flash(f'Sync (archivo) completo: {added} importados, {skipped} ya existían.', 'ok')
                return redirect(url_for('index'))

            # — Live sync from pWarehouse8 —
            import requests as _req
            url = PWAREHOUSE_URL.rstrip('/')
            rut = PWAREHOUSE_RUT
            pw  = PWAREHOUSE_PASS
            if not rut or not pw:
                flash('Configurá PWAREHOUSE_RUT y PWAREHOUSE_PASS en las variables de Railway.', 'err')
                return redirect(url_for('index'))

            sess = _req.Session()
            sess.headers['User-Agent'] = 'Mozilla/5.0 (compatible; Goodvalley-Sync/1.0)'
            r = sess.get(url + '/', timeout=30)
            r.raise_for_status()

            sid = None
            for pat in [
                r"['\"]?_S_ID['\"]?\s*[=:]\s*['\"]([A-Za-z0-9]+)['\"]",
                r"_S_ID=([A-Za-z0-9]+)",
            ]:
                m = _re.search(pat, r.text, _re.IGNORECASE)
                if m:
                    sid = m.group(1)
                    break
            if not sid:
                raise ValueError("No se encontró _S_ID en la página de pWarehouse8.")

            # _fp_ must be EMPTY; O16/O17 are sent as separate POST fields
            sess.post(url + '/HandleEvent', data={
                'Ajax': '1', 'IsEvent': '1', 'Obj': 'O23', 'Evt': 'click',
                'this': 'O23', '_S_ID': sid, '_fp_': '',
                'O16': ' \x02\x02' + rut,
                'O17': ' \x02\x02' + pw,
                '_seq_': 'a', '_uo_': 'O0',
            }, timeout=30)

            all_rows, start, total, page_size = [], 0, None, 2000
            while True:
                page = start // page_size + 1
                dr = sess.get(url + '/HandleEvent', params={
                    'IsEvent': '1', 'Obj': 'O16B', 'Evt': 'data',
                    'options': '1', 'page': str(page),
                    'start': str(start), 'limit': str(page_size), '_S_ID': sid,
                }, timeout=90)
                dr.raise_for_status()
                raw = dr.text.strip()
                if raw in ('{[]}', '{}', '[]', ''):
                    raise ValueError(
                        "pWarehouse8 bloqueó la conexión desde Railway (respondió vacío). "
                        "Ejecutá sync_local.py desde tu Mac y después subí el bins.json "
                        "con el botón '📤 Subir JSON'."
                    )
                data = _parse_js(raw)
                rows = data.get('rows', [])
                if not rows:
                    break
                all_rows.extend(rows)
                if total is None:
                    total = int(data.get('results', len(rows)))
                if start + page_size >= total:
                    break
                start += page_size

            bins_data = []
            for row in all_rows:
                producto = str(_rv(row, 8) or '').strip()
                if 'CIRUELA' not in producto.upper():
                    continue

                tarja = _rv(row, 1)
                if tarja is None:
                    continue
                tarja_str = str(int(float(tarja))) if isinstance(tarja, (int, float)) else str(tarja).strip()

                neto = _rv(row, 2)
                try:
                    weight = float(neto) if neto is not None else 0.0
                except Exception:
                    weight = 0.0

                # Caliber from SERIE col or PRODUCTO
                caliber = None
                serie = _rv(row, 15)
                if serie:
                    s = str(serie).strip()
                    import re as _reb
                    m = _reb.search(r'(\d{2,3}/\d{2,3}|\d{2,3}\+)', s)
                    if m:
                        caliber = m.group(1)

                # Drying from SECADO col or PRODUCTO
                secado = _rv(row, 16)
                drying = _DRYING_MAP.get(str(secado).lower().strip()) if secado else None

                # Fallback: parse both from PRODUCTO
                cal_p, dry_p = _parse_producto(producto)
                if not caliber:
                    caliber = cal_p
                if not drying:
                    drying = dry_p

                if not drying:
                    continue

                productor = _rv(row, 12)
                bins_data.append({
                    'bin_identifier': tarja_str,
                    'producto':       producto,
                    'caliber':        caliber or '',
                    'drying':         drying,
                    'weight_kg':      weight,
                    'humedad':        None,
                    'contenedor':     '',
                    'producer_name':  str(productor).strip() if productor else '',
                    'temporada':      _temporada(tarja_str),
                })

            added, skipped = _do_import(bins_data)
            flash(f'Sync en vivo completo: {added} importados, {skipped} ya existían.', 'ok')

        except Exception as e:
            db.session.rollback()
            flash(f'Sync falló: {e}', 'err')

        return redirect(url_for('index'))

    @app.route('/sync/upload', methods=['POST'])
    def sync_upload():
        import json as _json
        from models import Bin

        def _temporada(t):
            if len(t) >= 8:
                try:
                    p = int(t[:2])
                    if 18 <= p <= 35:
                        return str(2000 + p)
                except ValueError:
                    pass
            return None

        f = request.files.get('bins_file')
        if not f or not f.filename:
            flash('No se seleccionó ningún archivo.', 'err')
            return redirect(url_for('index'))

        try:
            bins_data = _json.load(f)
        except Exception as e:
            flash(f'El archivo no es JSON válido: {e}', 'err')
            return redirect(url_for('index'))

        try:
            from models import OrderLine, Order, Allocation
            existing_map = {
                row[0]: row[1] for row in
                db.session.query(Bin.bin_identifier, Bin.id)
                .filter(Bin.status == 'available').all()
            }
            all_ids = {
                row[0] for row in db.session.query(Bin.bin_identifier).all()
            }

            added = updated = skipped = 0
            new_batch = []
            # track (bin_identifier, ot) for new bins that carry OT info
            to_allocate = []

            for b in bins_data:
                bid = str(b.get('bin_identifier', '')).strip()
                if not bid:
                    skipped += 1
                    continue
                drying = b.get('drying') or ''
                if drying not in ('cancha', 'horno', 'termino_secado'):
                    skipped += 1
                    continue

                weight = float(b.get('weight_kg') or 0)

                if bid in existing_map:
                    db.session.query(Bin).filter_by(id=existing_map[bid]).update({
                        'weight_kg': weight,
                        'humedad': b.get('humedad'),
                        'caliber': b.get('caliber') or '',
                        'drying': drying,
                        'producto': b.get('producto') or '',
                        'contenedor': b.get('contenedor') or '',
                        'producer_name': b.get('producer_name') or '',
                        'temporada': b.get('temporada') or _temporada(bid),
                    }, synchronize_session=False)
                    updated += 1
                elif bid not in all_ids:
                    new_batch.append(Bin(
                        bin_identifier=bid,
                        producto=b.get('producto') or '',
                        caliber=b.get('caliber') or '',
                        drying=drying,
                        weight_kg=weight,
                        humedad=b.get('humedad'),
                        contenedor=b.get('contenedor') or '',
                        producer_name=b.get('producer_name') or '',
                        temporada=b.get('temporada') or _temporada(bid),
                        status='available',
                    ))
                    all_ids.add(bid)
                    added += 1
                    ot = b.get('ot')
                    if ot:
                        to_allocate.append((bid, str(ot).strip()))
                    if len(new_batch) >= 500:
                        db.session.bulk_save_objects(new_batch)
                        db.session.commit()
                        new_batch = []
                else:
                    skipped += 1

            if new_batch:
                db.session.bulk_save_objects(new_batch)
            db.session.commit()

            # Auto-allocate new bins whose OT matches an open order line
            alloc_count = 0
            if to_allocate:
                # Build OT → line map from open/confirmed orders
                ot_line_map = {}
                for ln in (OrderLine.query
                           .join(Order, OrderLine.order_id == Order.id)
                           .filter(Order.status.in_(['open', 'confirmed']))
                           .filter(OrderLine.notes.isnot(None))
                           .all()):
                    for part in (ln.notes or '').split('·'):
                        part = part.strip()
                        if part.startswith('OT '):
                            ot_key = part[3:].strip()
                            ot_line_map.setdefault(ot_key, ln)

                for bid, ot in to_allocate:
                    line = ot_line_map.get(ot)
                    if not line:
                        continue
                    b_obj = Bin.query.filter_by(
                        bin_identifier=bid, status='available').first()
                    if not b_obj:
                        continue
                    already = Allocation.query.filter_by(bin_id=b_obj.id).first()
                    if already:
                        continue
                    db.session.add(Allocation(
                        order_id=line.order_id, line_id=line.id, bin_id=b_obj.id))
                    b_obj.status = 'allocated'
                    alloc_count += 1
                db.session.commit()

            msg = f'Sync completo: {added} nuevos, {updated} actualizados, {skipped} omitidos'
            if alloc_count:
                msg += f', {alloc_count} auto-asignados a órdenes'
            flash(msg + '.', 'ok')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al importar: {e}', 'err')

        return redirect(url_for('index'))

    # ── Bins ──────────────────────────────────────────────────────────────────

    @app.route('/bins')
    def list_bins():
        from models import Bin, CALIBER_OPTIONS, DRYING_LABELS
        from sqlalchemy import func

        q_caliber  = request.args.get('caliber', '')
        q_drying   = request.args.get('drying', '')
        q_status   = request.args.get('status', '')
        q_temporada = request.args.get('temporada', '')
        q_grower   = request.args.get('grower', '')
        q_text     = request.args.get('q', '').strip()

        query = Bin.query

        if q_caliber:
            query = query.filter(Bin.caliber == q_caliber)
        if q_drying:
            query = query.filter(Bin.drying == q_drying)
        if q_status:
            query = query.filter(Bin.status == q_status)
        if q_temporada:
            query = query.filter(Bin.temporada == q_temporada)
        if q_grower:
            query = query.filter(Bin.producer_name == q_grower)
        if q_text:
            like = f'%{q_text}%'
            query = query.filter(
                db.or_(Bin.bin_identifier.ilike(like), Bin.producto.ilike(like))
            )

        total_count = query.count()
        total_kg    = db.session.query(func.sum(Bin.weight_kg)).filter(
            *([Bin.caliber == q_caliber]  if q_caliber  else []),
            *([Bin.drying == q_drying]    if q_drying   else []),
            *([Bin.status == q_status]    if q_status   else []),
            *([Bin.temporada == q_temporada] if q_temporada else []),
            *([Bin.producer_name == q_grower] if q_grower else []),
        ).scalar() or 0

        bins = query.order_by(Bin.bin_identifier).limit(500).all()

        # Filter options
        growers = [
            r[0] for r in
            db.session.query(Bin.producer_name)
            .filter(Bin.producer_name != '')
            .distinct()
            .order_by(Bin.producer_name)
            .all()
        ]
        temporadas = [
            r[0] for r in
            db.session.query(Bin.temporada)
            .filter(Bin.temporada.isnot(None))
            .distinct()
            .order_by(Bin.temporada)
            .all()
        ]

        return render_template('bins/list.html',
            bins=bins,
            total_count=total_count,
            total_kg=round(total_kg, 1),
            CALIBER_OPTIONS=CALIBER_OPTIONS,
            DRYING_LABELS=DRYING_LABELS,
            growers=growers,
            temporadas=temporadas,
            q_caliber=q_caliber, q_drying=q_drying, q_status=q_status,
            q_temporada=q_temporada, q_grower=q_grower, q_text=q_text,
        )

    # ── Orders ────────────────────────────────────────────────────────────────

    @app.route('/orders')
    def list_orders():
        from models import Order
        orders = Order.query.order_by(Order.created_at.desc()).all()
        return render_template('orders/list.html', orders=orders)

    @app.route('/orders', methods=['POST'])
    def create_order():
        from models import Order, OrderLine

        customer  = request.form.get('customer', '').strip()
        reference = request.form.get('reference', '').strip() or None
        notes     = request.form.get('notes', '').strip() or None

        if not customer:
            flash('El nombre del cliente es obligatorio.', 'err')
            return redirect(url_for('list_orders'))

        try:
            total_kg = float(request.form.get('target_kg', 0))
        except (ValueError, TypeError):
            total_kg = 0.0

        max_humedad   = request.form.get('max_humedad')   or None
        temporada     = request.form.get('temporada')     or None
        product_type  = request.form.get('product_type')  or None
        fruit_quality = request.form.get('fruit_quality') or None

        # Build calibre/drying breakdown from percentage rows
        calibers     = request.form.getlist('caliber[]')
        caliber_pcts = request.form.getlist('caliber_pct[]')
        dryings      = request.form.getlist('drying[]')
        drying_pcts  = request.form.getlist('drying_pct[]')

        cal_pairs = []
        for c, p in zip(calibers, caliber_pcts):
            try:
                pct = float(p)
                if pct > 0:
                    cal_pairs.append((c or None, pct))
            except (ValueError, TypeError):
                pass

        dry_pairs = []
        for d, p in zip(dryings, drying_pcts):
            try:
                pct = float(p)
                if pct > 0:
                    dry_pairs.append((d or None, pct))
            except (ValueError, TypeError):
                pass

        # Default: single unconstrained line
        if not cal_pairs:
            cal_pairs = [(None, 100.0)]
        if not dry_pairs:
            dry_pairs = [(None, 100.0)]

        # Auto-generate OT: YYMMDD-N (same format as pWarehouse historico)
        from datetime import date as _date
        today_prefix = _date.today().strftime('%y%m%d')
        existing_today = Order.query.filter(Order.ot.like(f'{today_prefix}-%')).count()
        ot = f'{today_prefix}-{existing_today + 1}'

        order = Order(customer=customer, reference=reference, notes=notes, ot=ot)
        db.session.add(order)
        db.session.flush()

        for cal, cal_pct in cal_pairs:
            for dry, dry_pct in dry_pairs:
                line_kg = round(total_kg * (cal_pct / 100.0) * (dry_pct / 100.0), 1)
                line = OrderLine(
                    order_id=order.id,
                    caliber=cal,
                    drying=dry,
                    target_kg=line_kg,
                    max_humedad=float(max_humedad) if max_humedad else None,
                    temporada=temporada,
                    product_type=product_type,
                    fruit_quality=fruit_quality,
                )
                db.session.add(line)

        db.session.commit()
        n_lines = len(cal_pairs) * len(dry_pairs)
        flash(f'Orden {order.ot} creada con {n_lines} línea{"s" if n_lines != 1 else ""}.', 'ok')
        return redirect(url_for('order_detail', order_id=order.id))

    @app.route('/orders/new')
    def new_order():
        from models import CALIBER_OPTIONS, DRYING_LABELS, get_yield
        # Build yield preview keyed by caliber string for the JS preview
        yield_table = {}
        for pt in ['tsc', 'tcc', 'ss', 'elliot', 'cn']:
            by_cal = {}
            for c in CALIBER_OPTIONS:
                r = get_yield(pt, None, c)
                if r:
                    by_cal[c] = round(r, 3)
            if by_cal:
                yield_table[pt] = by_cal
        return render_template('orders/new.html',
            CALIBER_OPTIONS=CALIBER_OPTIONS,
            DRYING_LABELS=DRYING_LABELS,
            YIELD_TABLE=yield_table)

    @app.route('/orders/<int:order_id>')
    def order_detail(order_id):
        from models import Order, Bin, CALIBER_OPTIONS, DRYING_LABELS, Allocation, Excedente

        order = Order.query.get_or_404(order_id)

        search_line_id = request.args.get('search_line', type=int)
        search_bins = []
        search_excedentes = []
        saldo_tarjas = set()
        caliber_f = u_lb_f = u_lb_lo_f = u_lb_hi_f = None
        if search_line_id and order.status in ('open', 'confirmed'):
            line = next((l for l in order.lines if l.id == search_line_id), None)
            if line:
                allocated_bin_ids = {
                    a.bin_id for a in
                    Allocation.query.filter(Allocation.bin_id.isnot(None)).all()
                }
                q = Bin.query.filter_by(status='available')

                # Detect serie-range caliber (e.g. "83-97" has dash but no slash)
                _is_serie_range = bool(
                    line.caliber and '-' in line.caliber and '/' not in line.caliber
                )

                if _is_serie_range:
                    try:
                        _lo_s, _hi_s = line.caliber.split('-', 1)
                        _def_lo, _def_hi = int(_lo_s), int(_hi_s)
                    except (ValueError, AttributeError):
                        _def_lo = _def_hi = None
                    caliber_f  = request.args.get('caliber_f') or ''
                    u_lb_lo_f  = request.args.get('u_lb_lo_f', type=int) or _def_lo
                    u_lb_hi_f  = request.args.get('u_lb_hi_f', type=int) or _def_hi
                    if caliber_f:
                        q = q.filter(Bin.caliber == caliber_f)
                    if u_lb_lo_f is not None:
                        q = q.filter(Bin.u_lb >= float(u_lb_lo_f))
                    if u_lb_hi_f is not None:
                        q = q.filter(Bin.u_lb <= float(u_lb_hi_f))
                else:
                    caliber_f = request.args.get('caliber_f') or line.caliber
                    u_lb_f    = request.args.get('u_lb_f', type=int)
                    if caliber_f:
                        q = q.filter(Bin.caliber == caliber_f)
                    if u_lb_f:
                        q = q.filter(Bin.u_lb == float(u_lb_f))
                if line.drying:
                    q = q.filter(Bin.drying == line.drying)
                if line.temporada:
                    q = q.filter(Bin.temporada == line.temporada)
                if line.max_humedad:
                    q = q.filter(
                        db.or_(Bin.humedad.is_(None), Bin.humedad <= line.max_humedad)
                    )
                if allocated_bin_ids:
                    q = q.filter(Bin.id.notin_(allocated_bin_ids))
                search_bins = q.order_by(Bin.u_lb.asc().nulls_last(), Bin.bin_identifier).limit(200).all()

                # Compute saldo tarjas so they can be sorted to the top
                from models import HistoricoMovimiento, WASTE_SERIES as _WS
                _prod = (HistoricoMovimiento.query
                    .filter(HistoricoMovimiento.movimiento.in_([
                        'INGRESO DESDE PROCESO', 'REPALETIZAJE', 'INGRESO REEMBALAJE'
                    ]))
                    .filter(~db.or_(
                        HistoricoMovimiento.serie.in_(_WS),
                        HistoricoMovimiento.serie.ilike('%DESCARTE%')
                    ))
                    .all())
                _cons = (HistoricoMovimiento.query
                    .filter(HistoricoMovimiento.movimiento.in_([
                        'EMBARQUE', 'EGRESO A REPALETIZAJE', 'EGRESO REEMBALAJE'
                    ]))
                    .all())
                _cons_t, _shipped_ots = set(), set()
                for _r in _cons:
                    if _r.tarja: _cons_t.add(_r.tarja.strip())
                    if _r.movimiento == 'EMBARQUE': _shipped_ots.add(_r.ot)
                for _r in _prod:
                    _t = (_r.tarja or '').strip()
                    if _t and _r.ot in _shipped_ots and _t not in _cons_t:
                        saldo_tarjas.add(_t)
                search_bins.sort(
                    key=lambda b: (0 if b.bin_identifier in saldo_tarjas else 1, b.u_lb or float('inf'), b.bin_identifier)
                )

                allocated_surplus_ids = {
                    a.surplus_id for a in
                    Allocation.query.filter(Allocation.surplus_id.isnot(None)).all()
                }
                eq = Excedente.query.filter_by(status='available')
                if caliber_f:
                    eq = eq.filter(Excedente.caliber == caliber_f)
                if line.drying:
                    eq = eq.filter(Excedente.drying == line.drying)
                if line.temporada:
                    eq = eq.filter(Excedente.temporada == line.temporada)
                if allocated_surplus_ids:
                    eq = eq.filter(Excedente.id.notin_(allocated_surplus_ids))
                search_excedentes = eq.order_by(Excedente.created_at.desc()).all()

        return render_template('orders/detail.html',
            order=order,
            search_bins=search_bins,
            search_excedentes=search_excedentes,
            search_line_id=search_line_id,
            u_lb_f=u_lb_f if search_line_id else None,
            u_lb_lo_f=u_lb_lo_f if search_line_id else None,
            u_lb_hi_f=u_lb_hi_f if search_line_id else None,
            saldo_tarjas=saldo_tarjas,
            CALIBER_OPTIONS=CALIBER_OPTIONS,
            DRYING_LABELS=DRYING_LABELS,
        )

    # ── Order download: PDF ───────────────────────────────────────────────────

    @app.route('/orders/<int:order_id>/download.pdf')
    def download_order_pdf(order_id):
        from models import Order
        import io
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable)
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm

        order = Order.query.get_or_404(order_id)

        PURPLE = colors.HexColor('#3b0764')
        GREEN  = colors.HexColor('#2d6a4f')
        LTGRAY = colors.HexColor('#f4f4f8')
        LAVEND = colors.HexColor('#ede9fe')
        GRAY   = colors.HexColor('#666666')

        # Styles
        h1     = ParagraphStyle('h1',  fontName='Helvetica-Bold', fontSize=20, textColor=PURPLE, spaceAfter=2)
        sub_st = ParagraphStyle('sub', fontName='Helvetica',       fontSize=9,  textColor=GRAY, spaceAfter=10)
        foot   = ParagraphStyle('ft',  fontName='Helvetica',       fontSize=7.5, textColor=colors.HexColor('#999999'))
        lstat  = ParagraphStyle('ls',  fontName='Helvetica',       fontSize=8.5, textColor=colors.HexColor('#444444'))
        # Cell styles for table content (Paragraph handles encoding + wrapping)
        hdr_c  = ParagraphStyle('hc',  fontName='Helvetica-Bold',  fontSize=8,   textColor=colors.white,   leading=10)
        cell_c = ParagraphStyle('cc',  fontName='Helvetica',       fontSize=8.5, textColor=colors.black,   leading=11)
        tarja  = ParagraphStyle('tc',  fontName='Helvetica-Bold',  fontSize=8.5, textColor=PURPLE,         leading=11)
        tot_c  = ParagraphStyle('tot', fontName='Helvetica-Bold',  fontSize=9,   textColor=GREEN,          leading=12)
        sec_h  = ParagraphStyle('sh',  fontName='Helvetica-Bold',  fontSize=9.5, textColor=colors.white)

        def P(text, st): return Paragraph(str(text or '-'), st)

        # A4 landscape: usable width = 29.7 - 3cm margins = 26.7cm
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        # Column widths (sum = 26.7cm)
        COL_W = [4*cm, 5*cm, 2.5*cm, 3*cm, 2.5*cm, 1.7*cm, 8*cm]

        story = []

        # ── Page header ──
        story.append(Paragraph('GOODVALLEY', h1))
        ot_label = order.ot or f'#{order.id}'
        title = f'OT {ot_label}  |  {order.customer}'
        if order.reference:
            title += f'  |  PO: {order.reference}'
        story.append(Paragraph(title, sub_st))
        story.append(HRFlowable(width='100%', thickness=1, color=PURPLE, spaceAfter=8))

        # Order meta table
        meta = [
            [P('Cliente', ParagraphStyle('ml', fontName='Helvetica-Bold', fontSize=9, textColor=PURPLE)),
             P(order.customer or '-', cell_c),
             P('Referencia', ParagraphStyle('ml2', fontName='Helvetica-Bold', fontSize=9, textColor=PURPLE)),
             P(order.reference or '-', cell_c)],
            [P('Estado',  ParagraphStyle('ml3', fontName='Helvetica-Bold', fontSize=9, textColor=PURPLE)),
             P(order.status_label, cell_c),
             P('Fecha',   ParagraphStyle('ml4', fontName='Helvetica-Bold', fontSize=9, textColor=PURPLE)),
             P(order.created_at.strftime('%d/%m/%Y'), cell_c)],
        ]
        meta_tbl = Table(meta, colWidths=[2.5*cm, 9*cm, 2.5*cm, 9*cm])
        meta_tbl.setStyle(TableStyle([
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 0.4*cm))

        # ── Per-line sections ──
        for line in order.lines:
            if not line.allocations:
                continue

            # Section header — spec_label already contains caliber + drying + product_type
            sec_label = f'LINEA: {line.spec_label}   |   {line.target_kg:,.1f} kg PT pedidos'
            hdr_tbl = Table([[P(sec_label, sec_h)]], colWidths=[26.7*cm])
            hdr_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, -1), PURPLE),
                ('LEFTPADDING',   (0, 0), (-1, -1), 10),
                ('TOPPADDING',    (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ]))
            story.append(Spacer(1, 0.35*cm))
            story.append(hdr_tbl)
            story.append(Spacer(1, 0.15*cm))

            # Build rows — all cells are Paragraph objects
            col_headers = ['TARJA', 'PRODUCTO', 'CALIBRE', 'SECADO', 'KG MP', 'HUM %', 'PRODUCTOR']
            rows = [[P(h, hdr_c) for h in col_headers]]
            total_kg = 0.0

            for a in line.allocations:
                if a.bin:
                    b = a.bin
                    cal = b.caliber or 'N/A'
                    if b.u_lb and b.u_lb > 0:
                        cal += f' / {int(b.u_lb)}'
                    rows.append([
                        P(b.bin_identifier, tarja),
                        P(b.producto, cell_c),
                        P(cal, cell_c),
                        P(b.drying_label, cell_c),
                        P(f'{b.weight_kg:,.1f}' if b.weight_kg else '-', cell_c),
                        P(f'{b.humedad:.1f}' if b.humedad is not None else '-', cell_c),
                        P(b.producer_name, cell_c),
                    ])
                    total_kg += b.weight_kg or 0
                elif a.surplus:
                    s = a.surplus
                    rows.append([
                        P(f'Exc. #{s.source_order_id}', tarja),
                        P(s.producto, cell_c),
                        P(s.caliber or 'N/A', cell_c),
                        P(s.drying_label, cell_c),
                        P(f'{s.weight_kg:,.1f}' if s.weight_kg else '-', cell_c),
                        P(str(s.boxes) if s.boxes is not None else '-', cell_c),
                        P('-', cell_c),
                    ])
                    total_kg += s.weight_kg or 0

            n = len(rows)  # includes header row
            # Total row
            rows.append([
                P('', cell_c), P('', cell_c), P('', cell_c), P('', cell_c),
                P(f'{total_kg:,.1f} kg', tot_c),
                P('', cell_c), P('', cell_c),
            ])

            style_cmds = [
                # Header row
                ('BACKGROUND',    (0, 0),  (-1, 0),      GREEN),
                ('TOPPADDING',    (0, 0),  (-1, 0),      5),
                ('BOTTOMPADDING', (0, 0),  (-1, 0),      5),
                # Data rows
                ('TOPPADDING',    (0, 1),  (-1, n),      4),
                ('BOTTOMPADDING', (0, 1),  (-1, n),      4),
                # Grid on data
                ('GRID',          (0, 0),  (-1, n-1),    0.3, colors.HexColor('#dddddd')),
                # Total row
                ('BACKGROUND',    (0, n),  (-1, n),      LAVEND),
                ('LINEABOVE',     (0, n),  (-1, n),      1.2, GREEN),
                ('ALIGN',         (4, 0),  (4, -1),      'RIGHT'),
                ('ALIGN',         (5, 0),  (5, -1),      'RIGHT'),
            ]
            # Alternating row tint (skip header row 0)
            for i in range(2, n, 2):
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), LTGRAY))

            bin_tbl = Table(rows, colWidths=COL_W, repeatRows=1)
            bin_tbl.setStyle(TableStyle(style_cmds))
            story.append(bin_tbl)

            story.append(Spacer(1, 0.15*cm))
            story.append(Paragraph(
                f'KG PT pedidos: <b>{line.target_kg:,.1f} kg</b>  |  '
                f'KG MP asignados: <b>{total_kg:,.1f} kg</b>',
                lstat
            ))

        story.append(Spacer(1, 1*cm))
        story.append(Paragraph(
            f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}', foot
        ))

        doc.build(story)
        buf.seek(0)
        safe_name = (order.customer or 'cliente').replace(' ', '_')
        ot_file = (order.ot or str(order.id)).replace('/', '-')
        return send_file(buf,
                         download_name=f'orden_{ot_file}_{safe_name}.pdf',
                         as_attachment=True,
                         mimetype='application/pdf')

    # ── Order download: Excel ─────────────────────────────────────────────────

    @app.route('/orders/<int:order_id>/download.xlsx')
    def download_order_xlsx(order_id):
        from models import Order, DRYING_LABELS as _DL
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        order = Order.query.get_or_404(order_id)

        def fill(argb):
            return PatternFill(fill_type='solid', fgColor=argb)

        def border():
            s = Side(style='thin', color='FFDDDDDD')
            return Border(left=s, right=s, top=s, bottom=s)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Orden'

        # ── Company banner ──
        ws.merge_cells('A1:G1')
        ws['A1'] = 'GOODVALLEY'
        ws['A1'].font       = Font(bold=True, color='FFFFFFFF', size=16)
        ws['A1'].fill       = fill('FF3B0764')
        ws['A1'].alignment  = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 30

        # ── Order meta ──
        info = [
            ('Orden',       f'#{order.id}'),
            ('Cliente',     order.customer or '—'),
            ('Referencia',  order.reference or '—'),
            ('Estado',      order.status_label),
            ('Fecha',       order.created_at.strftime('%d/%m/%Y')),
        ]
        for i, (label, val) in enumerate(info, start=2):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True, color='FF3B0764', size=10)
            ws[f'B{i}'] = val
            ws[f'B{i}'].font = Font(size=10)

        row = len(info) + 3  # blank row after meta

        # ── Per-line sections ──
        for line in order.lines:
            if not line.allocations:
                continue

            # Section header — spec_label already contains caliber + drying + product_type
            spec = f'LÍNEA: {line.spec_label}   —   {line.target_kg:,.1f} kg PT'
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = spec
            ws[f'A{row}'].font       = Font(bold=True, color='FFFFFFFF', size=10)
            ws[f'A{row}'].fill       = fill('FF3B0764')
            ws[f'A{row}'].alignment  = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 20
            row += 1

            # Column headers
            headers = ['TARJA', 'PRODUCTO', 'CALIBRE', 'SECADO', 'KG MP', 'HUM %', 'PRODUCTOR']
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col_idx, value=h)
                c.font      = Font(bold=True, color='FFFFFFFF', size=10)
                c.fill      = fill('FF2D6A4F')
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = border()
            ws.row_dimensions[row].height = 16
            row += 1

            total_kg  = 0.0
            data_start = row

            for a in line.allocations:
                if a.bin:
                    b = a.bin
                    cal = b.caliber or 'N/A'
                    if b.u_lb and b.u_lb > 0:
                        cal += f' · {int(b.u_lb)}'
                    data = [
                        b.bin_identifier or '—',
                        b.producto or '—',
                        cal,
                        b.drying_label or '—',
                        b.weight_kg,
                        b.humedad,
                        b.producer_name or '—',
                    ]
                    total_kg += b.weight_kg or 0
                elif a.surplus:
                    s = a.surplus
                    data = [
                        f'Excedente #{s.source_order_id}',
                        s.producto or '—',
                        s.caliber or 'N/A',
                        s.drying_label or '—',
                        s.weight_kg,
                        s.boxes,
                        '—',
                    ]
                    total_kg += s.weight_kg or 0
                else:
                    continue

                alt = (row - data_start) % 2 == 1
                for col_idx, val in enumerate(data, start=1):
                    c = ws.cell(row=row, column=col_idx, value=val)
                    c.border = border()
                    c.font   = Font(
                        size=10,
                        bold=(col_idx == 1),
                        color=('FF3B0764' if col_idx == 1 else 'FF000000')
                    )
                    c.fill = fill('FFF8F8F8') if alt else fill('FFFFFFFF')
                    if col_idx == 5:  # KG MP
                        c.number_format = '#,##0.0'
                        c.alignment     = Alignment(horizontal='right')
                    elif col_idx == 6:  # Hum%
                        c.number_format = '0.0'
                        c.alignment     = Alignment(horizontal='right')
                row += 1

            # Total row
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}']       = 'TOTAL KG MP'
            ws[f'A{row}'].font  = Font(bold=True, size=10, color='FF2D6A4F')
            ws[f'A{row}'].fill  = fill('FFEDE9FE')
            ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'E{row}']       = total_kg
            ws[f'E{row}'].font  = Font(bold=True, size=11, color='FF2D6A4F')
            ws[f'E{row}'].fill  = fill('FFEDE9FE')
            ws[f'E{row}'].number_format = '#,##0.0'
            ws[f'E{row}'].alignment     = Alignment(horizontal='right')
            for col_idx in range(6, 8):
                ws.cell(row=row, column=col_idx).fill = fill('FFEDE9FE')
            ws.row_dimensions[row].height = 18
            row += 2  # blank row between lines

        # Footer
        ws[f'A{row}'] = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws[f'A{row}'].font = Font(size=8, color='FF999999')

        # Column widths
        widths = {'A': 20, 'B': 16, 'C': 12, 'D': 16, 'E': 10, 'F': 9, 'G': 20}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = (order.customer or 'cliente').replace(' ', '_')
        ot_file = (order.ot or str(order.id)).replace('/', '-')
        return send_file(
            buf,
            download_name=f'orden_{ot_file}_{safe_name}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/orders/<int:order_id>/lines', methods=['POST'])
    def add_line(order_id):
        from models import Order, OrderLine

        order = Order.query.get_or_404(order_id)
        if order.status in ('fulfilled', 'cancelled'):
            flash('No se pueden agregar líneas a una orden cerrada.', 'err')
            return redirect(url_for('order_detail', order_id=order_id))

        caliber       = request.form.get('caliber')       or None
        drying        = request.form.get('drying')        or None
        target_kg     = request.form.get('target_kg',  '0')
        max_humedad   = request.form.get('max_humedad')   or None
        temporada     = request.form.get('temporada')     or None
        product_type  = request.form.get('product_type')  or None
        fruit_quality = request.form.get('fruit_quality') or None
        notes         = request.form.get('notes', '').strip() or None

        try:
            tkg = float(target_kg)
        except (ValueError, TypeError):
            tkg = 0.0

        line = OrderLine(
            order_id=order_id, caliber=caliber, drying=drying,
            target_kg=tkg,
            max_humedad=float(max_humedad) if max_humedad else None,
            temporada=temporada, product_type=product_type,
            fruit_quality=fruit_quality, notes=notes,
        )
        db.session.add(line)
        db.session.commit()
        flash('Línea agregada.', 'ok')
        return redirect(url_for('order_detail', order_id=order_id))

    @app.route('/orders/<int:order_id>/lines/<int:line_id>/allocate', methods=['POST'])
    def allocate_bins(order_id, line_id):
        from models import Order, OrderLine, Allocation, Bin, Excedente

        order = Order.query.get_or_404(order_id)
        line  = OrderLine.query.get_or_404(line_id)

        if order.status in ('fulfilled', 'cancelled'):
            flash('No se pueden asignar bins a una orden cerrada.', 'err')
            return redirect(url_for('order_detail', order_id=order_id))

        bin_ids     = request.form.getlist('bin_ids')
        surplus_ids = request.form.getlist('surplus_ids')

        if not bin_ids and not surplus_ids:
            flash('Seleccioná al menos un bin o excedente.', 'err')
            return redirect(url_for('order_detail', order_id=order_id,
                                    search_line=line_id))

        count = 0
        for bid in bin_ids:
            try:
                b = Bin.query.get(int(bid))
                if not b or b.status != 'available':
                    continue
                db.session.add(Allocation(order_id=order_id, line_id=line_id, bin_id=b.id))
                b.status = 'allocated'
                count += 1
            except IntegrityError:
                db.session.rollback()

        for sid in surplus_ids:
            s = Excedente.query.get(int(sid))
            if not s or s.status != 'available':
                continue
            db.session.add(Allocation(order_id=order_id, line_id=line_id, surplus_id=s.id))
            s.status = 'allocated'
            count += 1

        db.session.commit()
        flash(f'{count} elemento(s) asignado(s).', 'ok')
        return redirect(url_for('order_detail', order_id=order_id))

    @app.route('/allocations/<int:alloc_id>/release', methods=['POST'])
    def release_bin(alloc_id):
        from models import Allocation, Bin, Excedente

        alloc = Allocation.query.get_or_404(alloc_id)
        order_id = alloc.order_id
        if alloc.bin_id:
            b = Bin.query.get(alloc.bin_id)
            if b:
                b.status = 'available'
        elif alloc.surplus_id:
            s = Excedente.query.get(alloc.surplus_id)
            if s:
                s.status = 'available'
        db.session.delete(alloc)
        db.session.commit()
        flash('Liberado.', 'ok')
        return redirect(url_for('order_detail', order_id=order_id))

    @app.route('/reset', methods=['POST'])
    def reset_inventory():
        from models import Bin, Order, OrderLine, Allocation, Excedente

        if request.form.get('passcode') != '001083748':
            flash('Código incorrecto.', 'err')
            return redirect(url_for('index'))

        Allocation.query.delete(synchronize_session=False)
        Excedente.query.delete(synchronize_session=False)
        OrderLine.query.delete(synchronize_session=False)
        Order.query.delete(synchronize_session=False)
        Bin.query.delete(synchronize_session=False)
        db.session.commit()
        flash('Reset completo — bins, órdenes y asignaciones eliminados.', 'ok')
        return redirect(url_for('index'))

    @app.route('/reset-historico', methods=['POST'])
    def reset_historico():
        from models import HistoricoMovimiento, Proceso, OrdenDeVenta, AppSetting

        if request.form.get('passcode') != '001083748':
            flash('Código incorrecto.', 'err')
            return redirect(url_for('list_procesos'))

        OrdenDeVenta.query.delete(synchronize_session=False)
        Proceso.query.delete(synchronize_session=False)
        HistoricoMovimiento.query.delete(synchronize_session=False)
        db.session.commit()
        flash('Reset completo — procesos, embarques, descartes y saldos eliminados.', 'ok')
        return redirect(url_for('list_procesos'))

    @app.route('/orders/<int:order_id>/delete', methods=['POST'])
    def delete_order(order_id):
        from models import Order, Excedente

        order = Order.query.get_or_404(order_id)
        if order.status not in ('fulfilled', 'cancelled'):
            flash('Solo se pueden eliminar órdenes cumplidas o canceladas.', 'err')
            return redirect(url_for('order_detail', order_id=order_id))

        # Null out FK references from excedentes → order lines before cascade delete
        line_ids = [l.id for l in order.lines]
        if line_ids:
            Excedente.query.filter(
                Excedente.source_line_id.in_(line_ids)
            ).update({'source_line_id': None, 'source_order_id': None},
                     synchronize_session=False)
        Excedente.query.filter_by(source_order_id=order_id).update(
            {'source_order_id': None}, synchronize_session=False)

        db.session.delete(order)
        db.session.commit()
        flash(f'Orden #{order_id} eliminada.', 'ok')
        return redirect(url_for('list_orders'))

    @app.route('/orders/<int:order_id>/dispatch', methods=['GET', 'POST'])
    def dispatch_order(order_id):
        from models import Order, Bin, Allocation, Excedente, DRYING_LABELS

        order = Order.query.get_or_404(order_id)
        if order.status != 'confirmed':
            flash('Solo se pueden despachar órdenes confirmadas.', 'err')
            return redirect(url_for('order_detail', order_id=order_id))

        if request.method == 'POST':
            created = 0
            for line in order.lines:
                for alloc in line.allocations:
                    if not alloc.bin_id:
                        continue
                    exc_kg_str    = request.form.get(f'exc_kg_{alloc.id}', '').strip()
                    exc_boxes_str = request.form.get(f'exc_boxes_{alloc.id}', '').strip()
                    try:
                        exc_kg = float(exc_kg_str) if exc_kg_str else 0.0
                    except ValueError:
                        exc_kg = 0.0
                    try:
                        exc_boxes = int(exc_boxes_str) if exc_boxes_str else None
                    except ValueError:
                        exc_boxes = None

                    if exc_kg > 0:
                        s = Excedente(
                            source_order_id=order_id,
                            source_line_id=line.id,
                            source_bin_tarja=alloc.bin.bin_identifier,
                            caliber=line.caliber,
                            drying=line.drying,
                            temporada=line.temporada,
                            producto=line.spec_label,
                            weight_kg=exc_kg,
                            boxes=exc_boxes,
                            status='available',
                        )
                        db.session.add(s)
                        created += 1

            # Mark allocated bins/surplus as shipped
            allocs = Allocation.query.filter_by(order_id=order_id).all()
            for a in allocs:
                if a.bin_id:
                    b = Bin.query.get(a.bin_id)
                    if b:
                        b.status = 'shipped'
                elif a.surplus_id:
                    s = Excedente.query.get(a.surplus_id)
                    if s:
                        s.status = 'shipped'

            order.status = 'fulfilled'
            db.session.commit()

            msg = 'Orden cumplida — items despachados.'
            if created:
                msg += f' {created} excedente(s) registrado(s) como disponibles.'
            flash(msg, 'ok')
            return redirect(url_for('order_detail', order_id=order_id))

        return render_template('orders/dispatch.html',
            order=order, DRYING_LABELS=DRYING_LABELS)

    @app.route('/orders/<int:order_id>/status', methods=['POST'])
    def update_order_status(order_id):
        from models import Order, Bin, Allocation

        order      = Order.query.get_or_404(order_id)
        new_status = request.form.get('status')

        allowed = {
            'open':      ['confirmed', 'cancelled'],
            'confirmed': ['fulfilled', 'cancelled'],
            'fulfilled': [],
            'cancelled': [],
        }
        if new_status not in allowed.get(order.status, []):
            flash(f'Transición de estado no permitida.', 'err')
            return redirect(url_for('order_detail', order_id=order_id))

        if new_status == 'cancelled':
            allocs = Allocation.query.filter_by(order_id=order_id).all()
            for a in allocs:
                if a.bin_id:
                    b = Bin.query.get(a.bin_id)
                    if b:
                        b.status = 'available'
                elif a.surplus_id:
                    from models import Excedente
                    s = Excedente.query.get(a.surplus_id)
                    if s:
                        s.status = 'available'
                db.session.delete(a)
            flash('Orden cancelada — items liberados.', 'ok')
        elif new_status == 'fulfilled':
            allocs = Allocation.query.filter_by(order_id=order_id).all()
            for a in allocs:
                if a.bin_id:
                    b = Bin.query.get(a.bin_id)
                    if b:
                        b.status = 'shipped'
                elif a.surplus_id:
                    from models import Excedente
                    s = Excedente.query.get(a.surplus_id)
                    if s:
                        s.status = 'shipped'
            flash('Orden cumplida — items marcados como despachados.', 'ok')
        else:
            from models import ORDER_STATUS_LABELS
            flash(f'Estado actualizado a "{ORDER_STATUS_LABELS.get(new_status, new_status)}".', 'ok')

        order.status = new_status
        db.session.commit()
        return redirect(url_for('order_detail', order_id=order_id))

    @app.route('/admin/import-pallets', methods=['POST'])
    def admin_import_pallets():
        """
        Create Bin records from finished-pallet data (Pallets en bodega) and
        allocate each one to the correct order line.
        Payload: { passcode, pallets: [{tarja, ot, customer, caliber, drying,
                   product_type, weight_kg, producto, temporada}] }
        """
        from models import Order, OrderLine, Bin, Allocation
        from flask import jsonify

        payload = request.get_json(force=True, silent=True) or {}
        if payload.get('passcode') != '001083748':
            return jsonify({'error': 'unauthorized'}), 403

        pallets = payload.get('pallets', [])

        # Pre-load orders: customer (upper) → Order
        orders_by_customer = {
            o.customer.strip().upper(): o
            for o in Order.query.all()
        }

        # Pre-load existing bin identifiers
        existing_bins = {
            row[0] for row in db.session.query(Bin.bin_identifier).all()
        }

        added = skipped_dup = skipped_no_order = skipped_no_line = allocated = 0
        errors = []

        for p in pallets:
            tarja    = str(p.get('tarja') or '').strip()
            ot       = str(p.get('ot')     or '').strip()
            customer = str(p.get('customer') or '').strip()
            caliber  = p.get('caliber') or None
            drying   = p.get('drying')  or None
            pt       = p.get('product_type') or None
            kg       = float(p.get('weight_kg') or 0)
            produto  = p.get('producto') or ''
            temp     = p.get('temporada')
            temporada = str(int(temp)) if temp else None

            if not tarja:
                continue
            if tarja in existing_bins:
                skipped_dup += 1
                continue

            # Find order
            order = orders_by_customer.get(customer.upper())
            if not order:
                skipped_no_order += 1
                errors.append(f'No order for customer "{customer}" (tarja {tarja})')
                continue

            # Find matching line: notes contains OT, or caliber+pt match
            line = None
            for ln in order.lines:
                notes_ot = f'OT {ot}'
                if ln.notes and notes_ot in ln.notes:
                    line = ln
                    break
            if not line:
                # Fallback: match by caliber + product_type
                for ln in order.lines:
                    if ln.caliber == caliber and ln.product_type == pt:
                        line = ln
                        break
            if not line and order.lines:
                # Last resort: first line of the order
                line = order.lines[0]
            if not line:
                skipped_no_line += 1
                continue

            # Create bin
            b = Bin(
                bin_identifier=tarja,
                producto=produto,
                caliber=caliber,
                drying=drying,
                weight_kg=kg,
                temporada=temporada,
                status='available',
            )
            db.session.add(b)
            db.session.flush()
            existing_bins.add(tarja)
            added += 1

            # Allocate
            try:
                db.session.add(Allocation(
                    order_id=order.id, line_id=line.id, bin_id=b.id))
                b.status = 'allocated'
                allocated += 1
            except Exception as e:
                errors.append(f'Alloc error {tarja}: {e}')

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

        return jsonify({
            'added': added,
            'allocated': allocated,
            'skipped_duplicate': skipped_dup,
            'skipped_no_order': skipped_no_order,
            'skipped_no_line': skipped_no_line,
            'errors': errors[:20],
        })

    @app.route('/admin/import-orders', methods=['POST'])
    def admin_import_orders():
        from models import Order, OrderLine, Bin, Allocation
        from flask import jsonify

        payload = request.get_json(force=True, silent=True) or {}
        if payload.get('passcode') != '001083748':
            return jsonify({'error': 'unauthorized'}), 403

        orders_data = payload.get('orders', [])
        results = []
        errors = []

        existing_customers = {o.customer.strip().upper() for o in Order.query.all()}

        for od in orders_data:
            customer = (od.get('customer') or '').strip()
            if not customer:
                continue
            if customer.upper() in existing_customers:
                results.append({'customer': customer, 'skipped': True, 'reason': 'already exists'})
                continue

            order = Order(
                customer=customer,
                reference=od.get('reference') or None,
                notes=od.get('notes') or None,
                status='open',
            )
            db.session.add(order)
            db.session.flush()
            existing_customers.add(customer.upper())

            order_result = {'customer': customer, 'order_id': order.id, 'lines': []}

            for ld in od.get('lines', []):
                line = OrderLine(
                    order_id=order.id,
                    caliber=ld.get('caliber') or None,
                    drying=ld.get('drying') or None,
                    target_kg=float(ld.get('target_kg') or 0),
                    temporada=str(int(ld['temporada'])) if ld.get('temporada') else None,
                    product_type=ld.get('product_type') or None,
                    notes=ld.get('notes') or None,
                )
                db.session.add(line)
                db.session.flush()

                allocated = []
                not_found = []
                for bid in (ld.get('bin_identifiers') or []):
                    bid = str(bid).strip()
                    b = Bin.query.filter_by(bin_identifier=bid, status='available').first()
                    if b:
                        try:
                            db.session.add(Allocation(
                                order_id=order.id, line_id=line.id, bin_id=b.id))
                            b.status = 'allocated'
                            allocated.append(bid)
                        except Exception:
                            db.session.rollback()
                            not_found.append(bid)
                    else:
                        not_found.append(bid)

                order_result['lines'].append({
                    'line_id': line.id,
                    'caliber': line.caliber,
                    'drying': line.drying,
                    'product_type': line.product_type,
                    'target_kg': line.target_kg,
                    'allocated': len(allocated),
                    'not_found': len(not_found),
                })

            try:
                db.session.commit()
                results.append(order_result)
            except Exception as e:
                db.session.rollback()
                errors.append({'customer': customer, 'error': str(e)})

        return jsonify({'created': len([r for r in results if 'order_id' in r]),
                        'skipped': len([r for r in results if r.get('skipped')]),
                        'errors': errors,
                        'orders': results})

    # ── Procesos (from Histórico) ─────────────────────────────────────────────

    @app.route('/procesos')
    def list_procesos():
        from models import Proceso
        from collections import OrderedDict
        procesos = Proceso.query.order_by(Proceso.ot).all()
        last_imported = procesos[0].imported_at if procesos else None

        # Group sub-OTs (e.g. 785-0001-1 + 785-0001-2) under their base OT
        groups = OrderedDict()
        for p in procesos:
            groups.setdefault(_ot_base(p.ot), []).append(p)

        display_rows = []
        for base_ot, procs in groups.items():
            kg_in  = sum(p.kg_entrada or 0 for p in procs) or None
            kg_out = sum(p.kg_salida_bueno or 0 for p in procs) or None
            rend   = round(kg_out / kg_in * 100, 1) if kg_in else None
            bins   = sum(p.bins_entrada or 0 for p in procs) or None
            estados = {p.estado for p in procs}
            estado = ('embarcado' if 'embarcado' in estados
                      else 'procesado' if 'procesado' in estados
                      else 'en proceso')
            all_prods = set()
            for p in procs:
                for prod in (p.productores or '').split(', '):
                    if prod.strip():
                        all_prods.add(prod.strip())
            all_inicio = [p.fecha_inicio for p in procs if p.fecha_inicio]
            all_fin    = [p.fecha_fin    for p in procs if p.fecha_fin]
            secados = {s.strip() for p in procs for s in (p.secado or '').split(',') if s.strip()}
            display_rows.append({
                'base_ot':      base_ot,
                'procs':        procs,
                'n_lineas':     len(procs),
                'bins_entrada': bins,
                'kg_entrada':   kg_in,
                'kg_salida_bueno': kg_out,
                'rendimiento_pct': rend,
                'estado':       estado,
                'productores':  ', '.join(sorted(all_prods)),
                'fecha_inicio': min(all_inicio) if all_inicio else None,
                'fecha_fin':    max(all_fin)    if all_fin    else None,
                'tipoproceso':  procs[0].tipoproceso,
                'secado':       ', '.join(sorted(secados)),
                'temporada':    procs[0].temporada,
            })

        return render_template('procesos.html', display_rows=display_rows, last_imported=last_imported)

    @app.route('/procesos/<path:ot>')
    def proceso_detail(ot):
        from models import Proceso, HistoricoMovimiento, WASTE_SERIES
        from flask import redirect, abort

        base = _ot_base(ot)
        # If someone navigates to a sub-OT directly, redirect to its base
        if base != ot:
            return redirect(url_for('proceso_detail', ot=base))

        all_procs = Proceso.query.order_by(Proceso.ot).all()
        sub_procs = [p for p in all_procs if _ot_base(p.ot) == base]
        if not sub_procs:
            abort(404)

        def _mov_sort_key(m):
            # Chronological order; EGRESO A PROCESO wins tiebreaker on same timestamp
            dt  = m.fecha or datetime.min
            tie = 0 if (m.movimiento or '').upper().strip() == 'EGRESO A PROCESO' else 1
            return (dt, tie)

        movimientos_by_ot = {}
        for p in sub_procs:
            movs = (
                HistoricoMovimiento.query
                .filter_by(ot=p.ot)
                .all()
            )
            movimientos_by_ot[p.ot] = sorted(movs, key=_mov_sort_key)

        return render_template('proceso_detail.html',
                               base_ot=base,
                               sub_procs=sub_procs,
                               movimientos_by_ot=movimientos_by_ot)

    # ── Órdenes de Venta / Embarques (from Histórico) ────────────────────────

    @app.route('/ordenes-de-venta')
    def list_ordenes_de_venta():
        from models import OrdenDeVenta
        from collections import OrderedDict
        ordenes = OrdenDeVenta.query.order_by(OrdenDeVenta.fecha_primer_embarque.desc().nullslast()).all()
        last_imported = ordenes[0].imported_at if ordenes else None

        # Group sub-OTs under their base OT
        groups = OrderedDict()
        for o in ordenes:
            groups.setdefault(_ot_base(o.ot), []).append(o)

        display_rows = []
        for base_ot, ords in groups.items():
            kg = sum(o.kg_embarcado or 0 for o in ords) or None
            fechas_p = [o.fecha_primer_embarque for o in ords if o.fecha_primer_embarque]
            fechas_u = [o.fecha_ultimo_embarque for o in ords if o.fecha_ultimo_embarque]
            proceso = next((o.proceso for o in ords if o.proceso), None)
            kg_salida = sum(o.proceso.kg_salida_bueno or 0 for o in ords if o.proceso) or None
            display_rows.append({
                'base_ot':               base_ot,
                'ords':                  ords,
                'n_lineas':              len(ords),
                'cliente':               ords[0].cliente,
                'kg_embarcado':          kg,
                'kg_salida_bueno':       kg_salida,
                'fecha_primer_embarque': min(fechas_p) if fechas_p else None,
                'fecha_ultimo_embarque': max(fechas_u) if fechas_u else None,
                'proceso':               proceso,
            })

        return render_template('ordenes_de_venta.html',
                               display_rows=display_rows,
                               last_imported=last_imported,
                               total_ordenes=len(display_rows),
                               total_kg=sum(r['kg_embarcado'] or 0 for r in display_rows))

    @app.route('/ordenes-de-venta/<path:ot>')
    def orden_de_venta_detail(ot):
        from models import OrdenDeVenta, HistoricoMovimiento
        from flask import redirect, abort

        base = _ot_base(ot)
        if base != ot:
            return redirect(url_for('orden_de_venta_detail', ot=base))

        all_ords = OrdenDeVenta.query.order_by(OrdenDeVenta.ot).all()
        sub_ords = [o for o in all_ords if _ot_base(o.ot) == base]

        embarques_by_ot = {}
        for o in sub_ords:
            embarques_by_ot[o.ot] = (
                HistoricoMovimiento.query
                .filter_by(ot=o.ot, movimiento='EMBARQUE')
                .order_by(HistoricoMovimiento.fecha)
                .all()
            )

        return render_template('orden_de_venta_detail.html',
                               base_ot=base,
                               sub_ords=sub_ords,
                               embarques_by_ot=embarques_by_ot)

    # ── Manejo de Descartes ───────────────────────────────────────────────────

    @app.route('/descartes')
    def list_descartes():
        from models import HistoricoMovimiento, WASTE_SERIES

        rows = (HistoricoMovimiento.query
                .filter(HistoricoMovimiento.movimiento == 'INGRESO DESDE PROCESO')
                .filter(
                    db.or_(
                        HistoricoMovimiento.serie.in_(WASTE_SERIES),
                        HistoricoMovimiento.serie.ilike('%DESCARTE%')
                    )
                )
                .order_by(HistoricoMovimiento.fecha.desc())
                .all())

        tipos = sorted({(r.serie or '').upper().strip() for r in rows if r.serie})
        total_kg = sum(r.neto or 0 for r in rows)

        return render_template('descartes.html',
                               rows=rows,
                               tipos=tipos,
                               total_kg=total_kg,
                               total_rows=len(rows))

    # ── Manejo de Saldos ──────────────────────────────────────────────────────

    @app.route('/saldos')
    def list_saldos():
        from models import HistoricoMovimiento, WASTE_SERIES
        from collections import OrderedDict as _OD

        # All non-waste tarjas that came out of processing.
        # REPALETIZAJE and INGRESO REEMBALAJE are output tarjas from repackaging —
        # their originals were already consumed via EGRESO A REPALETIZAJE / EGRESO REEMBALAJE.
        produced = (HistoricoMovimiento.query
                    .filter(HistoricoMovimiento.movimiento.in_([
                        'INGRESO DESDE PROCESO', 'REPALETIZAJE', 'INGRESO REEMBALAJE'
                    ]))
                    .filter(
                        ~db.or_(
                            HistoricoMovimiento.serie.in_(WASTE_SERIES),
                            HistoricoMovimiento.serie.ilike('%DESCARTE%')
                        )
                    )
                    .order_by(HistoricoMovimiento.fecha)
                    .all())

        # Tarjas that were shipped or merged (EMBARQUE / EGRESO A REPALETIZAJE / EGRESO REEMBALAJE)
        consumed_rows = (HistoricoMovimiento.query
                         .filter(HistoricoMovimiento.movimiento.in_([
                             'EMBARQUE', 'EGRESO A REPALETIZAJE', 'EGRESO REEMBALAJE'
                         ]))
                         .all())

        # Build {exact_ot: {tarja, ...}} for consumed lookup
        # Also track which exact OTs have at least one real EMBARQUE
        consumed_by_ot = {}
        ots_with_embarque = set()
        for r in consumed_rows:
            if r.tarja:
                consumed_by_ot.setdefault(r.ot, set()).add(r.tarja.strip())
            if r.movimiento == 'EMBARQUE':
                ots_with_embarque.add(r.ot)

        # Group produced by base OT; find leftover (produced but not consumed)
        groups = _OD()
        for r in produced:
            groups.setdefault(_ot_base(r.ot), []).append(r)

        saldos = []
        for base_ot, prod_rows in groups.items():
            # If nothing was ever shipped for any sub-OT, it's not a saldo —
            # it's just an order that hasn't shipped yet at all
            sub_ots = {r.ot for r in prod_rows}
            if not any(ot in ots_with_embarque for ot in sub_ots):
                continue

            leftover = []
            for r in prod_rows:
                tarja = (r.tarja or '').strip()
                ot_consumed = consumed_by_ot.get(r.ot, set())
                if not tarja or tarja not in ot_consumed:
                    leftover.append(r)
            if leftover:
                total_kg = sum(r.neto or 0 for r in leftover)
                saldos.append({
                    'base_ot':  base_ot,
                    'tarjas':   sorted(leftover, key=lambda r: r.fecha or datetime.min),
                    'total_kg': total_kg,
                    'n_tarjas': len(leftover),
                })

        saldos.sort(key=lambda s: s['total_kg'], reverse=True)
        total_kg_all = sum(s['total_kg'] for s in saldos)

        q_init = request.args.get('q', '')

        # Collect every tarja already shown in the Histórico saldos section
        historico_tarjas = set()
        for s in saldos:
            for r in s['tarjas']:
                if r.tarja:
                    historico_tarjas.add(r.tarja.strip())

        from models import Pallet
        saldo_pallets = [
            p for p in (
                Pallet.query
                .filter(Pallet.ot.in_(list(ots_with_embarque)))
                .filter(Pallet.weight_kg > 0)
                .order_by(Pallet.ot, Pallet.tarja)
                .all()
            ) if p.tarja not in historico_tarjas
        ] if ots_with_embarque else []
        return render_template('saldos.html',
                               saldos=saldos,
                               total_kg=total_kg_all,
                               total_ots=len(saldos),
                               q_init=q_init,
                               saldo_pallets=saldo_pallets,)

    # ── Import Histórico ──────────────────────────────────────────────────────

    def _run_historico_import(excel_bytes):
        """Core import logic. Takes raw Excel bytes, returns (n_rows, n_procs, n_odvs)."""
        import pandas as _pd
        from io import BytesIO
        from models import HistoricoMovimiento, Proceso, OrdenDeVenta, WASTE_SERIES

        df = _pd.read_excel(BytesIO(excel_bytes), engine='openpyxl')

        def _v(val):
            try:
                if _pd.isna(val):
                    return None
            except Exception:
                pass
            return val

        def _str(val):
            v = _v(val)
            return str(v).strip() if v is not None else None

        def _int(val):
            v = _v(val)
            try:
                return int(v) if v is not None else None
            except Exception:
                return None

        def _float(val):
            v = _v(val)
            try:
                return float(v) if v is not None else None
            except Exception:
                return None

        # ── 1. Full-replace historico_movimientos ─────────────────────────
        HistoricoMovimiento.query.delete(synchronize_session=False)
        db.session.commit()

        records = []
        for _, row in df.iterrows():
            records.append({
                'idpsj':           _float(row.get('IDPSJ')),
                'item':            _int(row.get('ITEM')),
                'cdgproducto':     _int(row.get('CDGPRODUCTO')),
                'idtransaccion':   _float(row.get('IDTRANSACCION')),
                'cdgcontenedor':   _int(row.get('CDGCONTENEDOR')),
                'cdgmvmnt':        _int(row.get('CDGMVMNT')),
                'cdgclase':        _int(row.get('CDGCLASE')),
                'cdgbodega':       _float(row.get('CDGBODEGA')),
                'ot':              _str(row.get('OT')) or '',
                'idot':            _int(row.get('IDOT')),
                'linea':           _float(row.get('LINEA')),
                'tipo':            _str(row.get('TIPO')),
                'revision':        _int(row.get('REVISION')),
                'movimiento':      _str(row.get('MOVIMIENTO')),
                'tipomovimiento':  _str(row.get('TIPOMOVIMIENTO')),
                'sestado':         _str(row.get('SESTADO')),
                'estado':          _int(row.get('ESTADO')),
                'estadoitem':      _float(row.get('ESTADOITEM')),
                'sestadoitem':     _float(row.get('SESTADOITEM')),
                'fecha':           _v(row.get('FECHA')),
                'fechaproduccion': _v(row.get('FECHAPRODUCCION')),
                'horaproduccion':  _str(row.get('HORAPRODUCCION')),
                'tarja':           _str(row.get('TARJA')),
                'serie':           _str(row.get('SERIE')),
                'lote':            _str(row.get('LOTE')),
                'guia':            _float(row.get('GUIA')),
                'producto':        _str(row.get('PRODUCTO')),
                'temporada':       _float(row.get('TEMPORADA')),
                'neto':            _float(row.get('NETO')),
                'bruto':           _float(row.get('BRUTO')),
                'tara':            _float(row.get('TARA')),
                'taracontenedor':  _float(row.get('TARACONTENEDOR')),
                'unidades':        _int(row.get('UNIDADES')),
                'unidad':          _str(row.get('UNIDAD')),
                'u_lb':            _float(row.get('U_LB')),
                'u_lb1':           _float(row.get('U_LB1')),
                'u_lb2':           _float(row.get('U_LB2')),
                'u_lb3':           _float(row.get('U_LB3')),
                'u_lb4':           _float(row.get('U_LB4')),
                'tipoproceso':     _str(row.get('TIPOPROCESO')),
                'secado':          _str(row.get('SECADO')),
                'productor':       _str(row.get('PRODUCTOR')),
                'rutproductor':    _str(row.get('RUTPRODUCTOR')),
                'exportador':      _str(row.get('EXPORTADOR')),
                'rutexportador':   _str(row.get('RUTEXPORTADOR')),
                'cliente':         _str(row.get('CLIENTE')),
                'usr':             _str(row.get('USR')),
                'turno':           _int(row.get('TURNO')),
                'contenedor':      _str(row.get('CONTENEDOR')),
                'tipocontenedor':  _str(row.get('TIPOCONTENEDOR')),
                'bodega':          _str(row.get('BODEGA')),
                'humedad':         _float(row.get('HUMEDAD')),
                'preservante':     _float(row.get('PRESERVANTE')),
                'aceite':          _float(row.get('ACEITE')),
                'carozo_col':      _float(row.get('CAROZO')),
                'pallet_clase':    _str(row.get('PALLET_CLASE')),
                's_pallet_clase':  _str(row.get('S_PALLET_CLASE')),
                'pallet_estado_ot': _str(row.get('PALLET_ESTADO_OT')),
                's_pallet_estado_ot': _str(row.get('S_PALLET_ESTADO_OT')),
                'pallet_estado_vigente': _str(row.get('PALLET_ESTADO_VIGENTE')),
                's_pallet_estado_vigente': _str(row.get('S_PALLET_ESTADO_VIGENTE')),
                'presenciametales': _str(row.get('PRESENCIAMETALES')),
                's_presenciametales': _str(row.get('S_PRESENCIAMETALES')),
                'idbins2':         _str(row.get('IDBINS2')),
                'count_ticket':    _float(row.get('COUNT_TICKET')),
                'ticket_pesaje':   _float(row.get('TICKET_PESAJE')),
                'documentoreferencia': _str(row.get('DOCUMENTOREFERENCIA')),
                'observaciones':   _str(row.get('OBSERVACIONES')),
                'idoe':            _float(row.get('IDOE')),
                'idsb':            _float(row.get('IDSB')),
                'sb':              _float(row.get('SB')),
                'idreproceso':     _float(row.get('IDREPROCESO')),
                'idrepaletizaje':  _float(row.get('IDREPALETIZAJE')),
                'idreembalaje':    _float(row.get('IDREEMBALAJE')),
                'idreenvasado':    _float(row.get('IDREENVASADO')),
                'x':               _str(row.get('X')),
                'y':               _float(row.get('Y')),
                'z':               _float(row.get('Z')),
                'direccion':       _float(row.get('DIRECCION')),
            })

        from sqlalchemy import insert as _sa_insert
        db.session.execute(_sa_insert(HistoricoMovimiento), records)
        db.session.commit()
        n_rows = len(records)

        # ── 2. Rebuild Proceso summaries ──────────────────────────────────
        OrdenDeVenta.query.delete(synchronize_session=False)
        Proceso.query.delete(synchronize_session=False)
        db.session.commit()

        proc_in_movs  = {'EGRESO A PROCESO'}
        proc_out_movs = {'INGRESO DESDE PROCESO'}

        mov_col = df['MOVIMIENTO'].fillna('')
        ot_col  = df['OT'].fillna('').astype(str).str.strip()
        serie_up = df['SERIE'].fillna('').str.upper()

        proc_ots = sorted(set(ot_col[mov_col.isin(proc_in_movs)]))
        now = datetime.utcnow()
        proc_records = []

        for ot in proc_ots:
            mask = ot_col == ot
            in_m  = mask & mov_col.isin(proc_in_movs)
            out_m = mask & mov_col.isin(proc_out_movs)
            emb_m = mask & (mov_col == 'EMBARQUE')

            waste_mask = serie_up.isin(WASTE_SERIES) | serie_up.str.contains('DESCARTE', na=False)
            carozo_mask   = serie_up == 'CAROZO'
            contra_mask   = serie_up == 'CONTRAMUESTRA'
            descarte_mask = serie_up.str.contains('DESCARTE', na=False)

            neto_in   = df.loc[in_m,  'NETO'].dropna()
            neto_good = df.loc[out_m & ~waste_mask, 'NETO'].dropna()
            neto_car  = df.loc[out_m & carozo_mask,   'NETO'].dropna()
            neto_des  = df.loc[out_m & descarte_mask,  'NETO'].dropna()
            neto_con  = df.loc[out_m & contra_mask,    'NETO'].dropna()
            neto_emb  = df.loc[emb_m, 'NETO'].dropna()

            kg_entrada      = float(neto_in.sum())   if not neto_in.empty   else None
            kg_salida_bueno = float(neto_good.sum()) if not neto_good.empty else None
            kg_carozo       = float(neto_car.sum())  if not neto_car.empty  else None
            kg_descarte     = float(neto_des.sum())  if not neto_des.empty  else None
            kg_contramuestra = float(neto_con.sum()) if not neto_con.empty  else None
            kg_embarcado    = float(neto_emb.sum())  if not neto_emb.empty  else None

            rend = None
            if kg_entrada and kg_entrada > 0 and kg_salida_bueno is not None:
                rend = round(kg_salida_bueno / kg_entrada * 100, 1)

            tp_vals = df.loc[mask, 'TIPOPROCESO'].dropna()
            sc_vals = df.loc[mask, 'SECADO'].dropna().unique()
            te_vals = df.loc[mask, 'TEMPORADA'].dropna()
            pr_vals = df.loc[in_m,  'PRODUCTOR'].dropna()
            io_vals = df.loc[mask, 'IDOT'].dropna()
            fi_vals = df.loc[in_m,  'FECHA'].dropna()
            fo_vals = df.loc[out_m, 'FECHA'].dropna()

            tipoproceso = str(tp_vals.iloc[0]).strip() if not tp_vals.empty else None
            secado = ', '.join(str(s).strip() for s in sc_vals if s) or None
            temporada = str(int(te_vals.iloc[0])) if not te_vals.empty else None
            productores = ', '.join(sorted({str(p).strip() for p in pr_vals if p})) or None
            idot = int(io_vals.iloc[0]) if not io_vals.empty else None

            fecha_inicio = fi_vals.min().to_pydatetime() if not fi_vals.empty else None
            fecha_fin    = fo_vals.max().to_pydatetime() if not fo_vals.empty else None

            if out_m.any() and emb_m.any():
                estado = 'embarcado'
            elif out_m.any():
                estado = 'procesado'
            else:
                estado = 'en proceso'

            proc_records.append({
                'ot': ot, 'idot': idot, 'temporada': temporada,
                'tipoproceso': tipoproceso, 'secado': secado,
                'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin,
                'bins_entrada': int(in_m.sum()),
                'kg_entrada': kg_entrada, 'kg_salida_bueno': kg_salida_bueno,
                'kg_carozo': kg_carozo, 'kg_descarte': kg_descarte,
                'kg_contramuestra': kg_contramuestra, 'kg_embarcado': kg_embarcado,
                'rendimiento_pct': rend, 'productores': productores,
                'estado': estado, 'imported_at': now,
            })

        if proc_records:
            db.session.execute(_sa_insert(Proceso), proc_records)
        db.session.commit()

        # ── 3. Rebuild OrdenDeVenta summaries ─────────────────────────────

        emb_mask = mov_col == 'EMBARQUE'
        emb_ots  = sorted(set(ot_col[emb_mask]))
        proc_id_by_ot = {p.ot: p.id for p in Proceso.query.all()}
        odv_records = []

        for ot in emb_ots:
            m = emb_mask & (ot_col == ot)
            cl_vals = df.loc[m, 'CLIENTE'].dropna()
            se_vals = df.loc[m, 'SERIE'].dropna().unique()
            ne_vals = df.loc[m, 'NETO'].dropna()
            fe_vals = df.loc[m, 'FECHA'].dropna()
            tp_vals = df.loc[m, 'TIPOPROCESO'].dropna()
            te_vals = df.loc[m, 'TEMPORADA'].dropna()
            io_vals = df.loc[m, 'IDOT'].dropna()

            odv_records.append({
                'ot':     ot,
                'idot':   int(io_vals.iloc[0]) if not io_vals.empty else None,
                'temporada': str(int(te_vals.iloc[0])) if not te_vals.empty else None,
                'tipoproceso': str(tp_vals.iloc[0]).strip() if not tp_vals.empty else None,
                'cliente': str(cl_vals.iloc[0]).strip() if not cl_vals.empty else None,
                'calibres': ', '.join(str(s).strip() for s in se_vals if s) or None,
                'kg_embarcado': float(ne_vals.sum()) if not ne_vals.empty else None,
                'fecha_primer_embarque': fe_vals.min().to_pydatetime() if not fe_vals.empty else None,
                'fecha_ultimo_embarque': fe_vals.max().to_pydatetime() if not fe_vals.empty else None,
                'proceso_id': proc_id_by_ot.get(ot),
                'imported_at': now,
            })

        if odv_records:
            db.session.execute(_sa_insert(OrdenDeVenta), odv_records)
        db.session.commit()

        return n_rows, len(proc_records), len(odv_records)

    @app.route('/admin/import-historico', methods=['POST'])
    def import_historico():
        f = request.files.get('historico_file')
        if not f:
            flash('No se seleccionó archivo.', 'err')
            return redirect(url_for('list_procesos'))
        try:
            n_rows, n_procs, n_odvs = _run_historico_import(f.read())
        except Exception as e:
            flash(f'Error al importar: {e}', 'err')
            return redirect(url_for('list_procesos'))
        flash(
            f'Histórico importado: {n_rows} movimientos. '
            f'{n_procs} procesos y {n_odvs} órdenes de venta reconstruidos.',
            'ok',
        )
        return redirect(url_for('list_procesos'))

    @app.route('/api/import-historico', methods=['POST'])
    def api_import_historico():
        """Passcode-protected endpoint for Google Apps Script to POST a Historico Excel."""
        import threading as _threading
        passcode = (request.form.get('passcode') or
                    request.headers.get('X-Passcode') or '').strip()
        if passcode != '001083748':
            return {'error': 'unauthorized'}, 401
        f = request.files.get('historico_file')
        if not f:
            return {'error': 'no file'}, 400
        excel_bytes = f.read()
        _app = app._get_current_object() if hasattr(app, '_get_current_object') else app

        def _run():
            try:
                with _app.app_context():
                    _run_historico_import(excel_bytes)
            except Exception as e:
                _app.logger.error(f'api_import_historico background error: {e}')

        _threading.Thread(target=_run, daemon=True).start()
        return {'ok': True, 'status': 'import started in background'}, 202

    @app.route('/pallets')
    def list_pallets():
        from models import Pallet, HistoricoMovimiento
        shipped_ots = {
            r[0] for r in db.session.query(HistoricoMovimiento.ot)
            .filter(HistoricoMovimiento.movimiento == 'EMBARQUE').all()
        }
        pallets   = Pallet.query.order_by(Pallet.ot, Pallet.tarja).all()
        last_sync = pallets[0].synced_at if pallets else None
        return render_template('pallets.html',
            pallets=pallets,
            shipped_ots=shipped_ots,
            last_sync=last_sync,
        )

    # ── Rendimientos ──────────────────────────────────────────────────────────

    @app.route('/rendimientos')
    def rendimientos():
        from models import YieldOverride, _YIELD_TABLE, _FLAT_YIELD

        ROWS_DEF = [
            ('TSC',    'tsc',     [35, 45, 55, 65, 75, 85, 95, 110, 132]),
            ('SS',     'ss',      [35, 45, 55, 65, 75, 85, 95]),
            ('TCC',    'tcc',     [35, 45, 55, 65, 75, 85, 95, 110, 132]),
            ('Elliot', 'elliot',  [95, 110, 132]),
            ('CN',     'natural', [35, 45, 55, 65, 75, 85, 90, 95, 110, 132, 157]),
        ]

        overrides = {
            (o.tipo, o.caliber_num): o
            for o in YieldOverride.query.all()
        }

        rows = []
        for tipo_label, yield_tipo, calibers in ROWS_DEF:
            for cal_num in calibers:
                ov = overrides.get((yield_tipo, cal_num))
                base = _YIELD_TABLE.get((yield_tipo, cal_num), _FLAT_YIELD.get(yield_tipo, 0))
                rend = ov.rend_teorico if ov else base
                rows.append({
                    'tipo_label':  tipo_label,
                    'yield_tipo':  yield_tipo,
                    'caliber_num': cal_num,
                    'rend_teorico': rend,
                    'comentario':  (ov.comentario or '') if ov else '',
                    'row_id':      f'{yield_tipo}_{cal_num}',
                })

        return render_template('rendimientos.html', rows=rows)

    @app.route('/rendimientos/check', methods=['POST'])
    def rendimientos_check():
        from flask import jsonify
        data = request.get_json(force=True, silent=True) or {}
        return jsonify({'ok': data.get('passcode') == '001083748'})

    @app.route('/rendimientos/save', methods=['POST'])
    def rendimientos_save():
        from models import YieldOverride, load_yield_overrides

        if request.form.get('passcode') != '001083748':
            flash('Contraseña incorrecta.', 'err')
            return redirect(url_for('rendimientos'))

        SAVE_DEFS = [
            ('tsc',     [35, 45, 55, 65, 75, 85, 95, 110, 132]),
            ('ss',      [35, 45, 55, 65, 75, 85, 95]),
            ('tcc',     [35, 45, 55, 65, 75, 85, 95, 110, 132]),
            ('elliot',  [95, 110, 132]),
            ('natural', [35, 45, 55, 65, 75, 85, 90, 95, 110, 132, 157]),
        ]

        for yield_tipo, calibers in SAVE_DEFS:
            for cal in calibers:
                rend_str = request.form.get(f'rend_{yield_tipo}_{cal}', '').strip()
                com = request.form.get(f'com_{yield_tipo}_{cal}', '').strip() or None
                try:
                    rend = float(rend_str)
                except (ValueError, TypeError):
                    continue
                ov = YieldOverride.query.filter_by(tipo=yield_tipo, caliber_num=cal).first()
                if ov:
                    ov.rend_teorico = rend
                    ov.comentario   = com
                else:
                    db.session.add(YieldOverride(
                        tipo=yield_tipo, caliber_num=cal,
                        rend_teorico=rend, comentario=com,
                    ))

        db.session.commit()
        load_yield_overrides()
        flash('Rendimientos actualizados.', 'ok')
        return redirect(url_for('rendimientos'))

    @app.route('/simulador')
    def simulador():
        return render_template('simulador.html')

    # ── Debug: pWarehouse scraper diagnostics ─────────────────────────────────

    @app.route('/admin/debug/proc')
    def debug_proc():
        from flask import make_response, send_file
        from pathlib import Path as _P
        fmt = request.args.get('fmt', 'text')
        base = _P('/tmp/gv_scraper')

        if fmt == 'img':
            name = request.args.get('name', 'proc_before.png')
            p = base / name
            if p.exists() and p.suffix == '.png':
                return send_file(str(p), mimetype='image/png')
            return 'Not found', 404

        if fmt == 'json':
            p = base / 'procesos_sample.json'
            if p.exists():
                return make_response(p.read_text(), 200,
                    {'Content-Type': 'application/json; charset=utf-8'})
            return 'Not found', 404

        # Default: HTML page with body text + links to images
        body_file = base / 'proc_body.txt'
        body_txt = body_file.read_text() if body_file.exists() else '(no proc_body.txt yet — run a sync first)'
        import html as _html
        imgs = ['proc_before.png', 'proc_after_clear.png', 'proc_nav_fail.png']
        img_tags = ''.join(
            f'<p><strong>{n}</strong><br>'
            f'<img src="/admin/debug/proc?fmt=img&name={n}" style="max-width:100%;border:1px solid #ccc"></p>'
            for n in imgs if (base / n).exists()
        )
        sample_link = (
            '<p><a href="/admin/debug/proc?fmt=json">procesos_sample.json</a></p>'
            if (base / 'procesos_sample.json').exists() else ''
        )
        html = f'''<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Debug Procesos</title>
<style>body{{font-family:monospace;padding:16px;background:#111;color:#ddd}}
pre{{background:#1a1a2a;padding:12px;white-space:pre-wrap;word-break:break-all;font-size:11px}}</style>
</head><body>
<h2>pWarehouse Informe Procesos — diagnóstico</h2>
{sample_link}
{img_tags}
<h3>proc_body.txt</h3>
<pre>{_html.escape(body_txt)}</pre>
</body></html>'''
        return make_response(html, 200, {'Content-Type': 'text/html; charset=utf-8'})

    return app




def _rebuild_summaries(df, WASTE_SERIES):
    """Rebuild Proceso and OrdenDeVenta summary tables from a historico DataFrame."""
    from models import Proceso, OrdenDeVenta
    from sqlalchemy import insert as _sa_insert

    OrdenDeVenta.query.delete(synchronize_session=False)
    Proceso.query.delete(synchronize_session=False)
    db.session.commit()

    mov_col  = df['MOVIMIENTO'].fillna('')
    ot_col   = df['OT'].fillna('').astype(str).str.strip()
    serie_up = df['SERIE'].fillna('').str.upper()
    now      = datetime.utcnow()

    proc_ots = sorted(set(ot_col[mov_col.isin({'EGRESO A PROCESO'})]))
    proc_records = []

    for ot in proc_ots:
        mask  = ot_col == ot
        in_m  = mask & mov_col.isin({'EGRESO A PROCESO'})
        out_m = mask & mov_col.isin({'INGRESO DESDE PROCESO'})
        emb_m = mask & (mov_col == 'EMBARQUE')

        waste_mask   = serie_up.isin(WASTE_SERIES) | serie_up.str.contains('DESCARTE', na=False)
        carozo_mask  = serie_up == 'CAROZO'
        contra_mask  = serie_up == 'CONTRAMUESTRA'
        descarte_mask = serie_up.str.contains('DESCARTE', na=False)

        neto_in   = df.loc[in_m,  'NETO'].dropna()
        neto_good = df.loc[out_m & ~waste_mask, 'NETO'].dropna()
        neto_car  = df.loc[out_m & carozo_mask,   'NETO'].dropna()
        neto_des  = df.loc[out_m & descarte_mask,  'NETO'].dropna()
        neto_con  = df.loc[out_m & contra_mask,    'NETO'].dropna()
        neto_emb  = df.loc[emb_m, 'NETO'].dropna()

        kg_entrada       = float(neto_in.sum())   if not neto_in.empty   else None
        kg_salida_bueno  = float(neto_good.sum()) if not neto_good.empty else None
        kg_carozo        = float(neto_car.sum())  if not neto_car.empty  else None
        kg_descarte      = float(neto_des.sum())  if not neto_des.empty  else None
        kg_contramuestra = float(neto_con.sum())  if not neto_con.empty  else None
        kg_embarcado     = float(neto_emb.sum())  if not neto_emb.empty  else None

        rend = None
        if kg_entrada and kg_entrada > 0 and kg_salida_bueno is not None:
            rend = round(kg_salida_bueno / kg_entrada * 100, 1)

        tp_vals = df.loc[mask, 'TIPOPROCESO'].dropna()
        sc_vals = df.loc[mask, 'SECADO'].dropna().unique()
        te_vals = df.loc[mask, 'TEMPORADA'].dropna()
        pr_vals = df.loc[in_m,  'PRODUCTOR'].dropna()
        io_vals = df.loc[mask, 'IDOT'].dropna()
        fi_vals = df.loc[in_m,  'FECHA'].dropna()
        fo_vals = df.loc[out_m, 'FECHA'].dropna()

        if out_m.any() and emb_m.any(): estado = 'embarcado'
        elif out_m.any():               estado = 'procesado'
        else:                           estado = 'en proceso'

        proc_records.append({
            'ot': ot, 'idot': int(io_vals.iloc[0]) if not io_vals.empty else None,
            'temporada': str(int(te_vals.iloc[0])) if not te_vals.empty else None,
            'tipoproceso': str(tp_vals.iloc[0]).strip() if not tp_vals.empty else None,
            'secado': ', '.join(str(s).strip() for s in sc_vals if s) or None,
            'fecha_inicio': fi_vals.min().to_pydatetime() if not fi_vals.empty else None,
            'fecha_fin':    fo_vals.max().to_pydatetime() if not fo_vals.empty else None,
            'bins_entrada': int(in_m.sum()), 'kg_entrada': kg_entrada,
            'kg_salida_bueno': kg_salida_bueno, 'kg_carozo': kg_carozo,
            'kg_descarte': kg_descarte, 'kg_contramuestra': kg_contramuestra,
            'kg_embarcado': kg_embarcado, 'rendimiento_pct': rend,
            'productores': ', '.join(sorted({str(p).strip() for p in pr_vals if p})) or None,
            'estado': estado, 'imported_at': now,
        })

    if proc_records:
        db.session.execute(_sa_insert(Proceso), proc_records)
    db.session.commit()

    emb_ots = sorted(set(ot_col[mov_col == 'EMBARQUE']))
    proc_id_by_ot = {p.ot: p.id for p in Proceso.query.all()}
    odv_records = []

    for ot in emb_ots:
        m = (mov_col == 'EMBARQUE') & (ot_col == ot)
        cl_vals = df.loc[m, 'CLIENTE'].dropna()
        se_vals = df.loc[m, 'SERIE'].dropna().unique()
        ne_vals = df.loc[m, 'NETO'].dropna()
        fe_vals = df.loc[m, 'FECHA'].dropna()
        tp_vals = df.loc[m, 'TIPOPROCESO'].dropna()
        te_vals = df.loc[m, 'TEMPORADA'].dropna()
        io_vals = df.loc[m, 'IDOT'].dropna()
        odv_records.append({
            'ot': ot, 'idot': int(io_vals.iloc[0]) if not io_vals.empty else None,
            'temporada': str(int(te_vals.iloc[0])) if not te_vals.empty else None,
            'tipoproceso': str(tp_vals.iloc[0]).strip() if not tp_vals.empty else None,
            'cliente': str(cl_vals.iloc[0]).strip() if not cl_vals.empty else None,
            'calibres': ', '.join(str(s).strip() for s in se_vals if s) or None,
            'kg_embarcado': float(ne_vals.sum()) if not ne_vals.empty else None,
            'fecha_primer_embarque': fe_vals.min().to_pydatetime() if not fe_vals.empty else None,
            'fecha_ultimo_embarque': fe_vals.max().to_pydatetime() if not fe_vals.empty else None,
            'proceso_id': proc_id_by_ot.get(ot), 'imported_at': now,
        })

    if odv_records:
        db.session.execute(_sa_insert(OrdenDeVenta), odv_records)
    db.session.commit()




def _pre_migrate(db_obj):
    """Drop tables whose schema changed — runs before db.create_all()."""
    stmts = [
        'DROP TABLE IF EXISTS proceso_lineas CASCADE',
        # Drop old procesos table if it lacks the new bins_entrada column
        """DO $$
        BEGIN
          IF NOT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_name='procesos' AND column_name='bins_entrada'
          ) THEN
            DROP TABLE IF EXISTS procesos CASCADE;
          END IF;
        END $$""",
    ]
    with db_obj.engine.connect() as conn:
        for sql in stmts:
            try:
                conn.execute(db_obj.text(sql))
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass


def _migrate(db_obj):
    """Additive schema migrations — safe to run on every startup."""
    stmts = [
        # bins — new columns
        'ALTER TABLE bins ADD COLUMN IF NOT EXISTS u_lb FLOAT',
        'ALTER TABLE bins ADD COLUMN IF NOT EXISTS producto VARCHAR(200)',
        'ALTER TABLE bins ADD COLUMN IF NOT EXISTS caliber VARCHAR(20)',
        'ALTER TABLE bins ADD COLUMN IF NOT EXISTS drying VARCHAR(30)',
        'ALTER TABLE bins ADD COLUMN IF NOT EXISTS humedad FLOAT',
        'ALTER TABLE bins ADD COLUMN IF NOT EXISTS contenedor VARCHAR(100)',
        'ALTER TABLE bins ADD COLUMN IF NOT EXISTS temporada VARCHAR(10)',
        "ALTER TABLE bins ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'available'",
        # Old drying_method column has NOT NULL — make it nullable so new inserts work
        'ALTER TABLE bins ALTER COLUMN drying_method DROP NOT NULL',
        # orders — new columns
        'ALTER TABLE orders ADD COLUMN IF NOT EXISTS customer VARCHAR(200)',
        'ALTER TABLE orders ADD COLUMN IF NOT EXISTS reference VARCHAR(100)',
        # copy buyer_name → customer for old rows
        'UPDATE orders SET customer = buyer_name WHERE customer IS NULL AND buyer_name IS NOT NULL',
        # Drop NOT NULL on legacy order columns so new inserts don't fail
        'ALTER TABLE orders ALTER COLUMN buyer_name DROP NOT NULL',
        # status rename (draft→open, shipped/closed→fulfilled)
        "UPDATE orders SET status = 'open'      WHERE status = 'draft'",
        "UPDATE orders SET status = 'fulfilled' WHERE status IN ('shipped','closed')",
        # excedentes support
        'ALTER TABLE allocations ALTER COLUMN bin_id DROP NOT NULL',
        'ALTER TABLE allocations ADD COLUMN IF NOT EXISTS surplus_id INTEGER REFERENCES excedentes(id)',
        # product_type replaces pitted on order lines
        'ALTER TABLE order_lines ADD COLUMN IF NOT EXISTS product_type VARCHAR(20)',
        # track which source bin each excedente came from
        'ALTER TABLE excedentes ADD COLUMN IF NOT EXISTS source_bin_tarja VARCHAR(50)',
        # fruit quality tier on order lines
        'ALTER TABLE order_lines ADD COLUMN IF NOT EXISTS fruit_quality VARCHAR(20)',
        # orders — auto-generated OT number (YYMMDD-N)
        'ALTER TABLE orders ADD COLUMN IF NOT EXISTS ot VARCHAR(20)',
        # warehouse location columns on historico_movimientos (added 2026-07)
        'ALTER TABLE historico_movimientos ADD COLUMN IF NOT EXISTS x VARCHAR(10)',
        'ALTER TABLE historico_movimientos ADD COLUMN IF NOT EXISTS y FLOAT',
        'ALTER TABLE historico_movimientos ADD COLUMN IF NOT EXISTS z FLOAT',
        'ALTER TABLE historico_movimientos ADD COLUMN IF NOT EXISTS direccion FLOAT',
        # app-wide key-value settings table
        """CREATE TABLE IF NOT EXISTS app_settings (
            key   VARCHAR(80) PRIMARY KEY,
            value TEXT
        )""",
        'ALTER TABLE pallets ADD COLUMN IF NOT EXISTS pallet_estado_ot VARCHAR(5)',
        'ALTER TABLE pallets ADD COLUMN IF NOT EXISTS s_pallet_clase VARCHAR(30)',
        'ALTER TABLE pallets ADD COLUMN IF NOT EXISTS unidades INTEGER',
    ]
    with db_obj.engine.connect() as conn:
        for sql in stmts:
            try:
                conn.execute(db_obj.text(sql))
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass


app = create_app()

if __name__ == '__main__':
    app.run(debug=True)
